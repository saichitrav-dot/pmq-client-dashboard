import io
import re
import subprocess
import unicodedata
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules import task2_assessment


QUADRANT_COLORS = {
    "Deployable Candidates": "#10b981",
    "Progressing Candidates": "#7c3aed",
    "Basic Competency": "#0ea5e9",
    "Critical Intervention": "#ef4444",
}

QUADRANT_ORDER = [
    "Deployable Candidates",
    "Progressing Candidates",
    "Basic Competency",
    "Critical Intervention",
]

QUADRANT_ACTIONS = {
    "Deployable Candidates": "Students in the A or A+ performance band and ready for accelerated deployment conversations.",
    "Progressing Candidates": "Students in the B performance band who are building momentum and need continued mentoring.",
    "Basic Competency": "Students in the C performance band who have cleared the baseline but still need guided improvement.",
    "Critical Intervention": "Students in the F performance band who need immediate academic and operational intervention.",
}

GRADE_ORDER = ["A+", "A", "B", "C", "F"]

ASSESSMENT_COMPOSITE_WEIGHT = 0.50
ATTENDANCE_WEIGHT = 0.10
TRAINER_FEEDBACK_WEIGHT = 0.05
ASSIGNMENT_WEIGHT = 0.35
DEDICATION_THRESHOLD = 60.0
PERFORMANCE_THRESHOLD = 60.0

QUADRANT_BUCKETS = {
    "A+": "Deployable Candidates",
    "A": "Deployable Candidates",
    "B": "Progressing Candidates",
    "C": "Basic Competency",
    "F": "Critical Intervention",
}

JECRC_BATCH_COLORS = [
    "1D4ED8",
    "7C3AED",
    "0EA5E9",
    "10B981",
    "F59E0B",
]

PUBLISHED_REPORT_PATH = Path(__file__).resolve().parents[1] / "published_reports" / "pmq_client_dashboard.xlsx"
REPO_ROOT = Path(__file__).resolve().parents[2]
PUBLISHED_REPORT_REPO_PATH = PUBLISHED_REPORT_PATH.relative_to(REPO_ROOT).as_posix()


def normalize_name(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9\s]", " ", text.lower())
    return re.sub(r"\s+", " ", text).strip()


def normalize_superset_id(value):
    text = re.sub(r"\s+", "", str(value or "")).strip()
    if text.lower() in {"", "nan", "none", "null"}:
        return ""
    return text[:-2] if text.endswith(".0") else text


def classify_code_score(score):
    numeric_score = pd.to_numeric(score, errors="coerce")
    if pd.isna(numeric_score):
        return "Not Classified"
    numeric_score = float(numeric_score)
    if numeric_score >= 90:
        return "Excellent"
    if numeric_score >= 80:
        return "Good"
    if numeric_score >= 70:
        return "Needs Attention"
    return "Retest Required"


def score_to_grade(score):
    numeric_score = pd.to_numeric(score, errors="coerce")
    if pd.isna(numeric_score):
        return "N/A"
    numeric_score = float(numeric_score)
    if numeric_score >= 90:
        return "A+"
    if numeric_score >= 80:
        return "A"
    if numeric_score >= 70:
        return "B"
    if numeric_score >= 60:
        return "C"
    return "F"


def performance_bucket(score):
    return QUADRANT_BUCKETS.get(score_to_grade(score), "Critical Intervention")


def batch_sort_key(value):
    text = str(value or "").strip()
    match = re.search(r"\bbatch[\s\-_]*(\d+)\b", text, re.IGNORECASE)
    if match:
        return (0, int(match.group(1)), text.lower())
    return (1, 9999, text.lower())


def derive_jecrc_batch_from_language(value):
    language = str(value or "").strip().lower()
    if not language:
        return ""
    if "java" in language:
        return "Batch 1 - JAVA Batch"
    if "pytest" in language:
        return "Batch 2 - Pytest Batch"
    if "robot" in language:
        return "Batch 3 - Robot Python"
    if "playwright" in language:
        return "Batch 4 - Playwright"
    if ".net" in language or "dotnet" in language:
        return "Batch 5 - Dotnet"
    return ""


def weighted_average(values):
    weighted_sum = 0.0
    total_weight = 0.0
    for value, weight in values:
        numeric_value = pd.to_numeric(value, errors="coerce")
        if pd.isna(numeric_value):
            continue
        weighted_sum += float(numeric_value) * float(weight)
        total_weight += float(weight)
    if total_weight <= 0:
        return np.nan
    return round(weighted_sum / total_weight, 1)


def fill_missing_scores_from_stream(series, total_weeks):
    numeric_series = pd.to_numeric(series, errors="coerce").clip(0, 100)
    if total_weeks > 0:
        return numeric_series.fillna(0)
    return numeric_series


def dataframe_from_session(data):
    if isinstance(data, pd.DataFrame):
        return data.copy()
    return pd.DataFrame(data or [])


def find_first_column(df, candidates):
    lowered = {str(column).strip().lower(): column for column in df.columns}
    for candidate in candidates:
        match = lowered.get(candidate.strip().lower())
        if match is not None:
            return match
    return None


def first_non_empty(series):
    for value in series:
        if pd.notna(value) and str(value).strip():
            return value
    return ""


def week_sort_key(value):
    text = str(value or "").strip()
    match = re.search(r"(\d+)", text)
    if match:
        return (0, int(match.group(1)), text.lower())
    return (1, 9999, text.lower())


def track_sort_key(value):
    text = str(value or "").strip().lower()
    if "assessment" in text:
        return (0, text)
    if "assignment" in text:
        return (1, text)
    normalized = task2_assessment.normalize_evaluation_track(value)
    return (2, normalized)


def build_identity_key(superset_id, candidate_name):
    normalized_id = normalize_superset_id(superset_id)
    if normalized_id:
        return f"id::{normalized_id}"
    normalized_name = normalize_name(candidate_name)
    if normalized_name:
        return f"name::{normalized_name}"
    return ""


def add_identity_fields(df, candidate_name_col, superset_candidates):
    working = df.copy()
    working["Candidate Name"] = working[candidate_name_col].fillna("").astype(str).str.strip()
    superset_col = find_first_column(working, superset_candidates)
    if superset_col is None:
        working["Superset ID"] = ""
    else:
        working["Superset ID"] = working[superset_col].apply(normalize_superset_id)
    working["Match_Name"] = working["Candidate Name"].apply(normalize_name)
    working["Identity_Key"] = working.apply(lambda row: build_identity_key(row["Superset ID"], row["Candidate Name"]), axis=1)
    return working[working["Candidate Name"].ne("") & working["Identity_Key"].ne("")].copy()


def dedupe_identity_frame(df, metric_column, keep_columns):
    if df.empty:
        return pd.DataFrame(columns=keep_columns)
    working = df.copy()
    working["__metric"] = pd.to_numeric(working.get(metric_column, 0), errors="coerce").fillna(-1.0)
    working["__has_id"] = working["Superset ID"].ne("")
    working = working[working["Identity_Key"].ne("")].copy()
    deduped = (
        working.sort_values(by=["__has_id", "__metric", "Candidate Name"], ascending=[False, False, True])
        .drop_duplicates(subset=["Identity_Key"], keep="first")
        .reset_index(drop=True)
    )
    return deduped[keep_columns]


def build_combined_eval_frame():
    uploaded_eval_df = dataframe_from_session(st.session_state.get("eval_results", []))
    task1_shared_eval_df = dataframe_from_session(st.session_state.get("task1_shared_eval_results", []))
    notices = []
    eval_frames = [frame for frame in [uploaded_eval_df, task1_shared_eval_df] if not frame.empty]
    if not eval_frames:
        return pd.DataFrame(columns=task2_assessment.EVAL_COLUMNS), notices
    combined_eval_df = pd.concat(eval_frames, ignore_index=True)
    normalized_eval_df, eval_notices = task2_assessment.normalize_eval_dataframe(combined_eval_df)
    notices.extend(eval_notices)
    dedupe_cols = [
        column
        for column in ["Superset ID", "Candidate Name", "AssessmentWeek", "EvaluationTrack", "Score", "Status", "Worksheet"]
        if column in normalized_eval_df.columns
    ]
    if dedupe_cols:
        normalized_eval_df = normalized_eval_df.drop_duplicates(subset=dedupe_cols, keep="last").reset_index(drop=True)
    return normalized_eval_df, notices


def attendance_date_columns(df):
    date_columns = []
    blocked_tokens = {
        "candidate", "name", "college", "cohort", "batch", "status", "attendance",
        "present", "feedback", "score", "grade", "superset", "technology", "language",
    }
    for column in df.columns:
        text = str(column).strip()
        lowered = text.lower()
        if any(token in lowered for token in blocked_tokens):
            continue
        if re.match(r"^\d{4}-\d{2}-\d{2}", text) or re.match(r"^\d{1,2}-[a-zA-Z]{3}-\d{2,4}", text):
            date_columns.append(column)
            continue
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
        if pd.notna(parsed):
            date_columns.append(column)
    return date_columns


def build_code_frame(eval_results):
    df = dataframe_from_session(eval_results)
    empty_columns = [
        "Identity_Key",
        "Superset ID",
        "Match_Name",
        "Code Candidate Name",
        "Raw Code Score",
        "PrimaryGap",
        "FilesAnalyzed",
        "ExperienceProfile",
        "Classification",
        "Code Weeks",
    ]
    if df.empty:
        return pd.DataFrame(columns=empty_columns)

    name_col = find_first_column(df, ["Candidate Name", "Trainee", "Student Name"])
    score_col = find_first_column(df, ["Score", "Raw Score", "AI Score", "Code Score"])
    if name_col is None or score_col is None:
        return pd.DataFrame(columns=empty_columns)

    working = add_identity_fields(df, name_col, ["Superset ID", "SupersetID", "Student ID", "Candidate ID"])
    if working.empty:
        return pd.DataFrame(columns=empty_columns)
    status_col = find_first_column(working, ["Status"])
    working["Raw Code Score"] = pd.to_numeric(working[score_col], errors="coerce").clip(0, 100)
    if status_col is not None:
        evaluated_mask = working[status_col].astype(str).str.strip().eq("Evaluated")
        working["Raw Code Score"] = working["Raw Code Score"].where(evaluated_mask, np.nan)
    working["PrimaryGap"] = working.get("PrimaryGap", "").fillna("").astype(str)
    working["FilesAnalyzed"] = pd.to_numeric(working.get("FilesAnalyzed", 0), errors="coerce").fillna(0).astype(int)
    working["ExperienceProfile"] = working.get("ExperienceProfile", "").fillna("").astype(str)
    working["Classification"] = working["Raw Code Score"].apply(classify_code_score)
    week_col = find_first_column(working, ["AssessmentWeek", "Week", "WeekLabel"])
    working["AssessmentWeek"] = working[week_col].fillna("").astype(str).str.strip() if week_col else ""
    working["Evaluated Week"] = np.where(working["Raw Code Score"].notna(), working["AssessmentWeek"], "")
    working["Code Candidate Name"] = working["Candidate Name"]
    working = working.sort_values(by=["Raw Code Score", "FilesAnalyzed", "Code Candidate Name"], ascending=[False, False, True], na_position="last")

    aggregated = (
        working.groupby("Identity_Key", as_index=False)
        .agg(
            {
                "Superset ID": first_non_empty,
                "Match_Name": first_non_empty,
                "Code Candidate Name": first_non_empty,
                "Raw Code Score": "mean",
                "PrimaryGap": first_non_empty,
                "FilesAnalyzed": "max",
                "ExperienceProfile": first_non_empty,
                "Classification": first_non_empty,
                "Evaluated Week": lambda series: len({str(value).strip() for value in series if str(value).strip()}),
            }
        )
        .rename(columns={"Evaluated Week": "Code Weeks"})
    )
    aggregated["Raw Code Score"] = aggregated["Raw Code Score"].round(1)
    aggregated["Code Weeks"] = pd.to_numeric(aggregated["Code Weeks"], errors="coerce").fillna(0).astype(int)
    return aggregated[empty_columns]


def build_github_track_frame(
    eval_results,
    evaluation_track,
    score_column,
    week_column,
    candidate_column,
    batch_column,
    technology_column,
):
    df = dataframe_from_session(eval_results)
    empty_columns = [
        "Identity_Key",
        "Superset ID",
        "Match_Name",
        candidate_column,
        batch_column,
        technology_column,
        score_column,
        week_column,
    ]
    if df.empty:
        return pd.DataFrame(columns=empty_columns)

    if "EvaluationTrack" not in df.columns:
        df["EvaluationTrack"] = task2_assessment.GITHUB_ASSESSMENT_TRACK
    df["EvaluationTrack"] = df["EvaluationTrack"].apply(task2_assessment.normalize_evaluation_track)
    working = df[df["EvaluationTrack"] == task2_assessment.normalize_evaluation_track(evaluation_track)].copy()
    if working.empty:
        return pd.DataFrame(columns=empty_columns)

    name_col = find_first_column(working, ["Candidate Name", "Trainee", "Student Name"])
    score_col = find_first_column(working, ["Score", "Raw Score", "AI Score", "Code Score"])
    if name_col is None or score_col is None:
        return pd.DataFrame(columns=empty_columns)

    working = add_identity_fields(working, name_col, ["Superset ID", "SupersetID", "Student ID", "Candidate ID"])
    if working.empty:
        return pd.DataFrame(columns=empty_columns)

    status_col = find_first_column(working, ["Status"])
    working[score_column] = pd.to_numeric(working[score_col], errors="coerce").clip(0, 100)
    if status_col is not None:
        evaluated_mask = working[status_col].astype(str).str.strip().eq("Evaluated")
        working[score_column] = working[score_column].where(evaluated_mask, np.nan)

    week_col = find_first_column(working, ["AssessmentWeek", "Week", "WeekLabel"])
    assigned_batch_col = find_first_column(working, ["Assigned Batch", "Batch", "Batch Name"])
    declared_technology_col = find_first_column(working, ["DeclaredTechnology", "Technology", "Framework", "Programming Language"])
    working["AssessmentWeek"] = working[week_col].fillna("").astype(str).str.strip() if week_col else ""
    working[batch_column] = working[assigned_batch_col].fillna("").astype(str).str.strip() if assigned_batch_col else ""
    working[technology_column] = (
        working[declared_technology_col].fillna("").astype(str).str.strip() if declared_technology_col else ""
    )
    total_uploaded_weeks = len({str(value).strip() for value in working["AssessmentWeek"] if str(value).strip()})
    working[candidate_column] = working["Candidate Name"]
    working = working.sort_values(
        by=[score_column, candidate_column, batch_column, technology_column],
        ascending=[False, True, True, True],
        na_position="last",
    )

    aggregated = (
        working.groupby("Identity_Key", as_index=False)
        .agg(
            {
                "Superset ID": first_non_empty,
                "Match_Name": first_non_empty,
                candidate_column: first_non_empty,
                batch_column: first_non_empty,
                technology_column: first_non_empty,
                score_column: "mean",
            }
        )
    )
    aggregated[score_column] = aggregated[score_column].round(1)
    aggregated[week_column] = int(total_uploaded_weeks)
    return aggregated[empty_columns]


def build_dropout_override_frame(eval_results):
    df = dataframe_from_session(eval_results)
    empty_columns = [
        "Identity_Key",
        "Superset ID",
        "Match_Name",
        "Dropout Candidate Name",
        "Dropout Assigned Batch",
        "Dropout Status",
    ]
    if df.empty or "Status" not in df.columns:
        return pd.DataFrame(columns=empty_columns)

    working = df.copy()
    working["Status"] = working["Status"].apply(task2_assessment.normalize_assessment_status)
    working = working[working["Status"].eq("Drop out")].copy()
    if working.empty:
        return pd.DataFrame(columns=empty_columns)

    name_col = find_first_column(working, ["Candidate Name", "Trainee", "Student Name"])
    if name_col is None:
        return pd.DataFrame(columns=empty_columns)
    working = add_identity_fields(working, name_col, ["Superset ID", "SupersetID", "Student ID", "Candidate ID"])
    if working.empty:
        return pd.DataFrame(columns=empty_columns)

    assigned_batch_col = find_first_column(working, ["Assigned Batch", "Batch", "Batch Name"])
    working["Dropout Candidate Name"] = working["Candidate Name"]
    working["Dropout Assigned Batch"] = working[assigned_batch_col].fillna("").astype(str).str.strip() if assigned_batch_col else ""
    working["Dropout Status"] = "Drop out"
    aggregated = (
        working.groupby("Identity_Key", as_index=False)
        .agg(
            {
                "Superset ID": first_non_empty,
                "Match_Name": first_non_empty,
                "Dropout Candidate Name": first_non_empty,
                "Dropout Assigned Batch": first_non_empty,
                "Dropout Status": first_non_empty,
            }
        )
    )
    return aggregated[empty_columns]


def build_tracker_frame(tracker_data):
    empty_columns = [
        "Identity_Key",
        "Superset ID",
        "Match_Name",
        "Tracker Candidate Name",
        "Trainer Rating",
        "Assigned Batch",
        "Theory Score",
        "Trainer Feedback",
        "Persona",
        "Tracker College",
    ]
    df = dataframe_from_session(tracker_data)
    if df.empty:
        return pd.DataFrame(columns=empty_columns)

    name_col = find_first_column(df, ["Candidate Name", "Trainee", "Student Name"])
    if name_col is None:
        return pd.DataFrame(columns=empty_columns)

    working = add_identity_fields(df, name_col, ["Superset ID", "SupersetID", "Student ID", "Candidate ID"])
    if working.empty:
        return pd.DataFrame(columns=empty_columns)

    rating_cols = [column for column in working.columns if "rating" in str(column).lower()]
    if rating_cols:
        rating_frame = working[rating_cols].apply(pd.to_numeric, errors="coerce")
        working["Trainer Rating"] = rating_frame.mean(axis=1).fillna(0).clip(0, 5)
    else:
        working["Trainer Rating"] = 0.0

    batch_col = find_first_column(working, ["Assigned Batch", "Batch", "Batch Name"])
    theory_col = find_first_column(working, ["Numeric_Rating", "Theory Score", "Total Score", "totalScore"])
    feedback_col = find_first_column(working, ["Trainer Feedback", "Feedback", "Comments"])
    persona_col = find_first_column(working, ["Persona", "Persona Label", "Classification"])
    college_col = find_first_column(working, ["College", "College Name", "Institute"])

    working["Assigned Batch"] = working[batch_col].fillna("Unknown Batch").astype(str).str.strip() if batch_col else "Unknown Batch"
    if theory_col is None:
        working["Theory Score"] = np.nan
    else:
        values = pd.to_numeric(working[theory_col], errors="coerce")
        if values.max(skipna=True) <= 5:
            working["Theory Score"] = (values * 20).clip(0, 100)
        else:
            working["Theory Score"] = values.clip(0, 100)
    working["Trainer Feedback"] = working[feedback_col].fillna("").astype(str) if feedback_col else ""
    working["Persona"] = working[persona_col].fillna("Persona not available").astype(str) if persona_col else "Persona not available"
    working["Tracker College"] = working[college_col].fillna("Unknown College").astype(str).str.strip() if college_col else "Unknown College"
    working["Tracker Candidate Name"] = working["Candidate Name"]
    working = working.sort_values(by=["Trainer Rating", "Theory Score", "Tracker Candidate Name"], ascending=[False, False, True], na_position="last")

    aggregated = working.groupby("Identity_Key", as_index=False).agg(
        {
            "Superset ID": first_non_empty,
            "Match_Name": first_non_empty,
            "Tracker Candidate Name": first_non_empty,
            "Trainer Rating": "mean",
            "Assigned Batch": first_non_empty,
            "Theory Score": "mean",
            "Trainer Feedback": first_non_empty,
            "Persona": first_non_empty,
            "Tracker College": first_non_empty,
        }
    )
    aggregated["Trainer Rating"] = aggregated["Trainer Rating"].round(2)
    aggregated["Theory Score"] = aggregated["Theory Score"].round(1)
    return aggregated[empty_columns]


def build_feedback_frame(tracker_data):
    raw_tracker_df = dataframe_from_session(tracker_data)
    if raw_tracker_df.empty:
        return pd.DataFrame(
            columns=[
                "Identity_Key",
                "Superset ID",
                "Match_Name",
                "Tracker Candidate Name",
                "Trainer Rating",
                "Assigned Batch",
                "Theory Score",
                "Trainer Feedback",
                "Persona",
                "Tracker College",
            ]
        )
    subjectivity_df, _ = task2_assessment.split_tracker_sources(raw_tracker_df)
    return build_tracker_frame(subjectivity_df)


def build_top_brains_frame(tracker_data):
    raw_tracker_df = dataframe_from_session(tracker_data)
    empty_columns = [
        "Identity_Key",
        "Superset ID",
        "Match_Name",
        "Top Brains Candidate Name",
        "Top Brains Score",
        "Top Brains Weeks",
    ]
    if raw_tracker_df.empty:
        return pd.DataFrame(columns=empty_columns)

    _, top_brains_df = task2_assessment.split_tracker_sources(raw_tracker_df)
    if top_brains_df.empty:
        return pd.DataFrame(columns=empty_columns)

    combined_df = task2_assessment.build_combined_tracker_dataset(top_brains_df)
    plot_df = task2_assessment.aggregate_tracker_scope(combined_df) if not combined_df.empty else pd.DataFrame()
    if plot_df.empty:
        return pd.DataFrame(columns=empty_columns)

    week_map = task2_assessment.build_week_map(plot_df)
    total_score_columns = [group.get("total_score") for group in week_map.values() if group.get("total_score")]
    total_marks_columns = [group.get("total_marks") for group in week_map.values() if group.get("total_marks")]
    if not total_score_columns:
        return pd.DataFrame(columns=empty_columns)

    working = plot_df.copy()
    working["Top Brains TotalScore"] = task2_assessment.average_numeric_columns(working, total_score_columns)
    working["Top Brains TotalMarks"] = task2_assessment.average_numeric_columns(working, total_marks_columns).replace(0, 100).fillna(100)
    working["Top Brains Score"] = task2_assessment.calculate_assessment_scores(
        working,
        None,
        "Top Brains TotalScore",
        "Top Brains TotalMarks",
    )
    week_presence = []
    for column in total_score_columns:
        week_presence.append(pd.to_numeric(working[column], errors="coerce").fillna(0).gt(0))
    working["Top Brains Weeks"] = pd.concat(week_presence, axis=1).sum(axis=1) if week_presence else 0
    working = add_identity_fields(working, "Candidate Name", ["Superset ID", "SupersetID", "Student ID", "Candidate ID"])
    if working.empty:
        return pd.DataFrame(columns=empty_columns)
    working["Top Brains Candidate Name"] = working["Candidate Name"]
    aggregated = (
        working.groupby("Identity_Key", as_index=False)
        .agg(
            {
                "Superset ID": first_non_empty,
                "Match_Name": first_non_empty,
                "Top Brains Candidate Name": first_non_empty,
                "Top Brains Score": "mean",
                "Top Brains Weeks": "max",
            }
        )
    )
    aggregated["Top Brains Score"] = aggregated["Top Brains Score"].round(1)
    aggregated["Top Brains Weeks"] = pd.to_numeric(aggregated["Top Brains Weeks"], errors="coerce").fillna(0).astype(int)
    return aggregated[empty_columns]


def compute_consistency_index(df):
    date_cols = attendance_date_columns(df)
    if not date_cols:
        return pd.Series(pd.to_numeric(df.get("Clean_Present_%", 0), errors="coerce").fillna(0).clip(0, 100), index=df.index)

    presence_frame = df[date_cols].astype(str).apply(
        lambda column: column.str.strip().str.upper().map({"P": 1, "A": 0, "L": 0, "OD": 1})
    )
    presence_frame = presence_frame.where(presence_frame.isin([0, 1]))
    variability = presence_frame.std(axis=1, skipna=True).fillna(0)
    attended_days = presence_frame.mean(axis=1, skipna=True).fillna(0) * 100
    consistency = (100 - (variability * 100)).clip(0, 100)
    return ((consistency * 0.6) + (attended_days * 0.4)).round(1)


def build_attendance_frame(att_data):
    empty_columns = [
        "Identity_Key",
        "Superset ID",
        "Match_Name",
        "Attendance Candidate Name",
        "Attendance %",
        "Consistency Index",
        "Standard_Status",
        "College",
        "Cohort",
    ]
    df = dataframe_from_session(att_data)
    if df.empty:
        return pd.DataFrame(columns=empty_columns)

    name_col = find_first_column(df, ["Candidate Name", "Trainee", "Student Name"])
    if name_col is None:
        return pd.DataFrame(columns=empty_columns)

    working = add_identity_fields(df, name_col, ["Superset ID", "SupersetID", "Student ID", "Candidate ID"])
    if working.empty:
        return pd.DataFrame(columns=empty_columns)

    attendance_col = find_first_column(working, ["Clean_Present_%", "Attendance %", "Present %", "% Present"])
    college_col = find_first_column(working, ["College", "College Name", "Institute"])
    status_col = find_first_column(working, ["Standard_Status", "Status", "Student Status"])
    cohort_col = find_first_column(working, ["Cohort", "Cohort Name"])

    if attendance_col:
        attendance_values = pd.to_numeric(working[attendance_col], errors="coerce").fillna(0)
        if attendance_values.max() <= 1.5:
            attendance_values = attendance_values * 100
        working["Attendance %"] = attendance_values.clip(0, 100)
    else:
        working["Attendance %"] = 0.0
    working["College"] = working[college_col].fillna("Unknown College").astype(str).str.strip() if college_col else "Unknown College"
    working["Standard_Status"] = working[status_col].fillna("Unknown").astype(str).str.strip() if status_col else "Unknown"
    working["Cohort"] = working[cohort_col].fillna("Unknown Cohort").astype(str).str.strip() if cohort_col else "Unknown Cohort"
    working["Consistency Index"] = compute_consistency_index(working)
    working["Attendance Candidate Name"] = working["Candidate Name"]
    working = working.sort_values(by=["Attendance %", "Attendance Candidate Name"], ascending=[False, True])

    aggregated = working.groupby("Identity_Key", as_index=False).agg(
        {
            "Superset ID": first_non_empty,
            "Match_Name": first_non_empty,
            "Attendance Candidate Name": first_non_empty,
            "Attendance %": "mean",
            "Consistency Index": "mean",
            "Standard_Status": first_non_empty,
            "College": first_non_empty,
            "Cohort": first_non_empty,
        }
    )
    aggregated["Attendance %"] = aggregated["Attendance %"].round(1)
    aggregated["Consistency Index"] = aggregated["Consistency Index"].round(1)
    return aggregated[empty_columns]


def merge_support_data(base_df, support_df, label):
    result = base_df.copy()
    match_col = f"{label} Match Type"
    result[match_col] = "Unmatched"
    support_columns = [column for column in support_df.columns if column not in {"Identity_Key", "Superset ID", "Match_Name"}]

    if support_df.empty:
        for column in support_columns:
            result[column] = np.nan
        return result

    support_by_id = support_df[support_df["Superset ID"].ne("")].drop_duplicates(subset=["Superset ID"], keep="first").set_index("Superset ID")
    unique_name_matches = {}
    for match_name, group in support_df[support_df["Match_Name"].ne("")].groupby("Match_Name"):
        unique_rows = group.drop_duplicates(subset=["Identity_Key"])
        if len(unique_rows) == 1:
            unique_name_matches[match_name] = unique_rows.iloc[0]

    row_updates = []
    for _, row in result.iterrows():
        matched_row = None
        match_type = "Unmatched"
        base_id = normalize_superset_id(row.get("Superset ID", ""))
        match_name = row.get("Match_Name", "")

        if base_id and base_id in support_by_id.index:
            matched_row = support_by_id.loc[base_id]
            if isinstance(matched_row, pd.DataFrame):
                matched_row = matched_row.iloc[0]
            match_type = "Superset ID"
        elif match_name in unique_name_matches:
            candidate = unique_name_matches[match_name]
            candidate_id = normalize_superset_id(candidate.get("Superset ID", ""))
            if not base_id or not candidate_id or candidate_id == base_id:
                matched_row = candidate
                match_type = "Name Fallback"

        update = {match_col: match_type}
        for column in support_columns:
            update[column] = matched_row.get(column, np.nan) if matched_row is not None else np.nan
        row_updates.append(update)

    updates_df = pd.DataFrame(row_updates, index=result.index)
    for column in updates_df.columns:
        result[column] = updates_df[column]
    return result


def build_executive_frame(
    assessment_github_df,
    assignment_github_df,
    feedback_df,
    top_brains_df,
    attendance_df,
    dropout_override_df,
    population_scope,
):
    attendance_base = attendance_df.copy()
    if attendance_base.empty:
        return pd.DataFrame()

    if dropout_override_df is not None and not dropout_override_df.empty:
        attendance_base = merge_support_data(attendance_base, dropout_override_df, "Dropout Override")
        empty_text = pd.Series(index=attendance_base.index, dtype=object)
        attendance_base["Standard_Status"] = (
            attendance_base.get("Dropout Status", empty_text)
            .replace("", np.nan)
            .fillna(attendance_base.get("Standard_Status", empty_text))
            .fillna("Unknown")
        )
        attendance_base["Assigned Batch"] = (
            attendance_base.get("Dropout Assigned Batch", empty_text)
            .replace("", np.nan)
            .fillna(attendance_base.get("Assigned Batch", empty_text))
            .fillna("Unknown Batch")
        )

    if population_scope == "Active Only":
        active_mask = attendance_base["Standard_Status"].astype(str).str.contains("Active", case=False, na=False)
        if active_mask.any():
            attendance_base = attendance_base.loc[active_mask].copy()

    if attendance_base.empty:
        return pd.DataFrame()

    assessment_total_weeks = 0
    if not assessment_github_df.empty and "Assessment GitHub Weeks" in assessment_github_df.columns:
        assessment_total_weeks = int(pd.to_numeric(assessment_github_df["Assessment GitHub Weeks"], errors="coerce").fillna(0).max() or 0)

    assignment_total_weeks = 0
    if not assignment_github_df.empty and "Assignment GitHub Weeks" in assignment_github_df.columns:
        assignment_total_weeks = int(pd.to_numeric(assignment_github_df["Assignment GitHub Weeks"], errors="coerce").fillna(0).max() or 0)

    top_brains_total_weeks = 0
    if not top_brains_df.empty and "Top Brains Weeks" in top_brains_df.columns:
        top_brains_total_weeks = int(pd.to_numeric(top_brains_df["Top Brains Weeks"], errors="coerce").fillna(0).max() or 0)

    merged_df = merge_support_data(attendance_base, feedback_df, "Feedback")
    merged_df = merge_support_data(merged_df, top_brains_df, "Top Brains")
    merged_df = merge_support_data(merged_df, assessment_github_df, "Assessment GitHub")
    merged_df = merge_support_data(merged_df, assignment_github_df, "Assignment GitHub")
    empty_text = pd.Series(index=merged_df.index, dtype=object)
    merged_df["Candidate Name"] = (
        merged_df["Attendance Candidate Name"]
        .fillna(merged_df.get("Tracker Candidate Name", empty_text))
        .fillna(merged_df.get("Top Brains Candidate Name", empty_text))
        .fillna(merged_df.get("Assessment GitHub Candidate Name", empty_text))
        .fillna(merged_df.get("Assignment GitHub Candidate Name", empty_text))
    )
    merged_df["College"] = merged_df["College"].fillna(merged_df.get("Tracker College", empty_text)).fillna("Unknown College")
    merged_df["Assigned Batch"] = (
        merged_df.get("Assessment GitHub Assigned Batch", empty_text)
        .replace("", np.nan)
        .fillna(merged_df.get("Assignment GitHub Assigned Batch", empty_text).replace("", np.nan))
        .fillna(merged_df.get("Assigned Batch", empty_text))
        .fillna("Unknown Batch")
    )
    merged_df["Programming Language"] = (
        merged_df.get("Assessment GitHub Technology", empty_text)
        .replace("", np.nan)
        .fillna(merged_df.get("Assignment GitHub Technology", empty_text).replace("", np.nan))
        .fillna(merged_df.get("DeclaredTechnology", empty_text))
        .fillna("Not Declared")
    )
    jecrc_mask = merged_df["College"].fillna("").astype(str).str.contains("JECRC", case=False, na=False)
    github_batch_missing_mask = (
        merged_df.get("Assessment GitHub Assigned Batch", empty_text).fillna("").astype(str).str.strip().eq("")
        & merged_df.get("Assignment GitHub Assigned Batch", empty_text).fillna("").astype(str).str.strip().eq("")
    )
    generic_batch_mask = merged_df["Assigned Batch"].fillna("").astype(str).str.strip().str.fullmatch(r"Batch\s*\d+", case=False)
    pending_mask = jecrc_mask & github_batch_missing_mask & generic_batch_mask
    derived_batch_series = merged_df["Programming Language"].apply(derive_jecrc_batch_from_language)
    merged_df.loc[pending_mask & derived_batch_series.ne(""), "Assigned Batch"] = derived_batch_series[pending_mask & derived_batch_series.ne("")]
    merged_df.loc[pending_mask & derived_batch_series.eq(""), "Assigned Batch"] = "Batch Mapping Pending"
    merged_df["Trainer Rating"] = pd.to_numeric(merged_df.get("Trainer Rating", 0), errors="coerce").fillna(0).clip(0, 5)
    merged_df["Trainer Feedback"] = merged_df.get("Trainer Feedback", empty_text).fillna("").astype(str)
    merged_df["Trainer Feedback Score"] = (merged_df["Trainer Rating"] * 20).clip(0, 100)
    merged_df["Assessment GitHub Score"] = fill_missing_scores_from_stream(merged_df.get("Assessment GitHub Score", np.nan), assessment_total_weeks)
    merged_df["Assignment GitHub Score"] = fill_missing_scores_from_stream(merged_df.get("Assignment GitHub Score", np.nan), assignment_total_weeks)
    merged_df["Top Brains Score"] = fill_missing_scores_from_stream(merged_df.get("Top Brains Score", np.nan), top_brains_total_weeks)
    merged_df["Assessment Composite Score"] = merged_df.apply(
        lambda row: weighted_average(
            [
                (row.get("Assessment GitHub Score"), 1.0),
                (row.get("Top Brains Score"), 1.0),
            ]
        ),
        axis=1,
    )
    merged_df["Performance Score"] = merged_df.apply(
        lambda row: weighted_average(
            [
                (row.get("Assessment Composite Score"), ASSESSMENT_COMPOSITE_WEIGHT),
                (row.get("Assignment GitHub Score"), ASSIGNMENT_WEIGHT),
            ]
        ),
        axis=1,
    )
    merged_df["Dedication Score"] = merged_df.apply(
        lambda row: weighted_average(
            [
                (row.get("Attendance %"), ATTENDANCE_WEIGHT),
                (row.get("Trainer Feedback Score"), TRAINER_FEEDBACK_WEIGHT),
            ]
        ),
        axis=1,
    )
    merged_df["Overall Performance Score"] = merged_df.apply(
        lambda row: weighted_average(
            [
                (row.get("Assessment Composite Score"), ASSESSMENT_COMPOSITE_WEIGHT),
                (row.get("Assignment GitHub Score"), ASSIGNMENT_WEIGHT),
                (row.get("Attendance %"), ATTENDANCE_WEIGHT),
                (row.get("Trainer Feedback Score"), TRAINER_FEEDBACK_WEIGHT),
            ]
        ),
        axis=1,
    )
    merged_df["Assessment Composite Grade"] = merged_df["Assessment Composite Score"].apply(score_to_grade)
    merged_df["Assignment Grade"] = merged_df["Assignment GitHub Score"].apply(score_to_grade)
    merged_df["Attendance Grade"] = merged_df["Attendance %"].apply(score_to_grade)
    merged_df["Trainer Feedback Grade"] = merged_df["Trainer Feedback Score"].apply(score_to_grade)
    merged_df["Performance Grade"] = merged_df["Performance Score"].apply(score_to_grade)
    merged_df["Dedication Grade"] = merged_df["Dedication Score"].apply(score_to_grade)
    merged_df["Overall Grade"] = merged_df["Overall Performance Score"].apply(score_to_grade)
    merged_df["Persona"] = merged_df.get("Persona", empty_text).fillna("Persona not available")
    merged_df["Assessment GitHub Weeks"] = pd.to_numeric(merged_df.get("Assessment GitHub Weeks", 0), errors="coerce").fillna(assessment_total_weeks).astype(int)
    merged_df["Assignment GitHub Weeks"] = pd.to_numeric(merged_df.get("Assignment GitHub Weeks", 0), errors="coerce").fillna(assignment_total_weeks).astype(int)
    merged_df["Top Brains Weeks"] = pd.to_numeric(merged_df.get("Top Brains Weeks", 0), errors="coerce").fillna(top_brains_total_weeks).astype(int)
    merged_df["Superset ID"] = merged_df["Superset ID"].apply(normalize_superset_id)
    return merged_df[merged_df["Candidate Name"].astype(str).str.strip().ne("")].copy()


def assign_quadrant(row):
    performance_score = pd.to_numeric(row.get("Performance Score"), errors="coerce")
    performance_score = 0.0 if pd.isna(performance_score) else float(performance_score)
    return performance_bucket(performance_score)


def build_data_quality_metrics(df):
    fallback_series = []
    for column in ["Feedback Match Type", "Top Brains Match Type", "Assessment GitHub Match Type", "Assignment GitHub Match Type"]:
        if column in df.columns:
            fallback_series.append(df[column].eq("Name Fallback"))
    fallback_count = int(pd.concat(fallback_series, axis=1).any(axis=1).sum()) if fallback_series else 0
    return {
        "Population": int(len(df)),
        "Feedback ID Matches": int(df.get("Feedback Match Type", pd.Series(index=df.index, dtype=object)).eq("Superset ID").sum()),
        "Top Brains ID Matches": int(df.get("Top Brains Match Type", pd.Series(index=df.index, dtype=object)).eq("Superset ID").sum()),
        "Assessment GitHub ID Matches": int(df.get("Assessment GitHub Match Type", pd.Series(index=df.index, dtype=object)).eq("Superset ID").sum()),
        "Assignment GitHub ID Matches": int(df.get("Assignment GitHub Match Type", pd.Series(index=df.index, dtype=object)).eq("Superset ID").sum()),
        "Name Fallbacks": fallback_count,
        "Missing Feedback Signal": int(df.get("Feedback Match Type", pd.Series(index=df.index, dtype=object)).eq("Unmatched").sum()),
        "Missing Top Brains Signal": int(df.get("Top Brains Match Type", pd.Series(index=df.index, dtype=object)).eq("Unmatched").sum()),
        "Missing Assessment GitHub": int(df.get("Assessment GitHub Match Type", pd.Series(index=df.index, dtype=object)).eq("Unmatched").sum()),
        "Missing Assignment GitHub": int(df.get("Assignment GitHub Match Type", pd.Series(index=df.index, dtype=object)).eq("Unmatched").sum()),
    }


def build_exec_insights(df, quality_metrics, population_scope):
    if df.empty:
        return []

    total = len(df)
    counts = df["Quadrant"].value_counts()
    avg_overall = pd.to_numeric(df["Overall Performance Score"], errors="coerce").mean()
    avg_performance = pd.to_numeric(df["Performance Score"], errors="coerce").mean()
    deployable_share = counts.get("Deployable Candidates", 0) / total
    progressing_share = counts.get("Progressing Candidates", 0) / total
    basic_share = counts.get("Basic Competency", 0) / total
    critical_share = counts.get("Critical Intervention", 0) / total

    insights = [
        {
            "headline": f"{deployable_share:.0%} of the {population_scope.lower()} cohort are deployment-ready.",
            "detail": "These students are landing in the A or A+ performance band after the weighted assessment-composite and assignment calculations.",
        },
        {
            "headline": f"{progressing_share:.0%} are in the B-grade progressing band.",
            "detail": "They are moving in the right direction and need continued assignment practice plus assessment reinforcement to move into the deployable band.",
        },
        {
            "headline": f"{basic_share:.0%} are in the C-grade basic-competency band.",
            "detail": "These students have crossed the baseline but still need structured technical improvement before they are management-ready for deployment review.",
        },
    ]

    if critical_share > 0:
        insights.append(
            {
                "headline": f"{critical_share:.0%} are in the F-grade critical-intervention band.",
                "detail": "They should move into the highest-priority remediation track with close trainer follow-up and targeted performance recovery.",
            }
        )
    elif avg_performance < PERFORMANCE_THRESHOLD:
        insights.append(
            {
                "headline": f"Average performance score is {avg_performance:.1f}, so the technical cohort average is still below the comfort band.",
                "detail": "Assessment composite and assignment GitHub outcomes need stronger uplift before more students can move into the deployable band.",
            }
        )
    else:
        insights.append(
            {
                "headline": f"Average overall performance score is {avg_overall:.1f}, which is currently stable.",
                "detail": "The executive score remains available for leadership reviews, while the category bucket is now driven only by the performance score grade band.",
            }
        )

    if quality_metrics["Name Fallbacks"] > 0:
        insights.append(
            {
                "headline": f"{quality_metrics['Name Fallbacks']} row(s) still required name fallback after the Superset-ID pass.",
                "detail": "Those rows are still usable, but they should be reviewed as data-quality exceptions rather than treated as ideal keyed merges.",
            }
        )
    return insights[:4]


def build_action_roster(df):
    roster = df.copy()
    return roster[
        [
            "Superset ID",
            "Candidate Name",
            "College",
        "Assigned Batch",
        "Programming Language",
        "Cohort",
            "Standard_Status",
            "Attendance %",
            "Attendance Grade",
            "Trainer Feedback Score",
            "Trainer Feedback Grade",
            "Assessment GitHub Score",
            "Assessment GitHub Weeks",
            "Top Brains Score",
            "Top Brains Weeks",
            "Assessment Composite Score",
            "Assessment Composite Grade",
            "Assignment GitHub Score",
            "Assignment GitHub Weeks",
            "Assignment Grade",
            "Performance Score",
            "Performance Grade",
            "Overall Performance Score",
            "Overall Grade",
            "Quadrant",
        ]
    ].copy()


def build_action_roster_display(df):
    roster = df.copy()
    display_columns = [
        "Superset ID",
        "Candidate Name",
        "College",
        "Assigned Batch",
        "Programming Language",
        "Standard_Status",
        "Attendance %",
        "Attendance Grade",
        "Trainer Feedback Score",
        "Trainer Feedback Grade",
        "Assessment GitHub Score",
        "Assessment GitHub Weeks",
        "Top Brains Score",
        "Top Brains Weeks",
        "Assessment Composite Score",
        "Assessment Composite Grade",
        "Assignment GitHub Score",
        "Assignment GitHub Weeks",
        "Assignment Grade",
        "Performance Score",
        "Performance Grade",
        "Overall Performance Score",
        "Overall Grade",
    ]
    return roster[[column for column in display_columns if column in roster.columns]].copy()


def to_excel_bytes(df):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="executive_roster")
    buffer.seek(0)
    return buffer.getvalue()


def publish_dashboard_report_file(report_bytes: bytes) -> Path:
    PUBLISHED_REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    PUBLISHED_REPORT_PATH.write_bytes(report_bytes)
    return PUBLISHED_REPORT_PATH


def run_git_command(args):
    return subprocess.run(
        ["git", *args],
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )


def push_published_dashboard_report():
    if not (REPO_ROOT / ".git").exists():
        return False, "Git repository was not found for the published dashboard."

    branch_result = run_git_command(["branch", "--show-current"])
    current_branch = branch_result.stdout.strip() if branch_result.returncode == 0 else ""
    if not current_branch:
        return False, "Unable to detect the current Git branch for dashboard publishing."

    diff_result = run_git_command(["diff", "--quiet", "--", PUBLISHED_REPORT_REPO_PATH])
    if diff_result.returncode not in {0, 1}:
        return False, (diff_result.stderr or diff_result.stdout or "Unable to inspect workbook changes before push.").strip()

    if diff_result.returncode == 0:
        push_result = run_git_command(["push", "origin", current_branch])
        if push_result.returncode != 0:
            return False, (push_result.stderr or push_result.stdout or "Git push failed.").strip()
        return True, f"Published dashboard workbook is already current on GitHub branch `{current_branch}`."

    commit_message = "Update published PMQ dashboard workbook"
    commit_result = run_git_command(["commit", "--only", PUBLISHED_REPORT_REPO_PATH, "-m", commit_message])
    if commit_result.returncode != 0:
        return False, (commit_result.stderr or commit_result.stdout or "Git commit failed for the published workbook.").strip()

    push_result = run_git_command(["push", "origin", current_branch])
    if push_result.returncode != 0:
        return False, (push_result.stderr or push_result.stdout or "Git push failed for the published workbook.").strip()

    commit_hash_result = run_git_command(["rev-parse", "--short", "HEAD"])
    commit_hash = commit_hash_result.stdout.strip() if commit_hash_result.returncode == 0 else current_branch
    return True, f"Published dashboard workbook pushed to GitHub on `{current_branch}` at commit `{commit_hash}`."


def build_jecrc_weekly_bifurcation_frames(filtered_df, eval_df):
    empty_batch_weekly = pd.DataFrame(columns=["Assigned Batch", "Assessment Week", "Average Score", "Submission Count"])
    empty_track_weekly = pd.DataFrame(columns=["Track", "Assessment Week", "Average Score", "Submission Count"])
    empty_batch_summary = pd.DataFrame(
        columns=["Assigned Batch", "Assessments Done", "Assignments Done", "Average Trainer Feedback Score", "Average Weekly Assessment Score"]
    )

    jecrc_df = filtered_df.copy()
    if "College" in jecrc_df.columns:
        jecrc_df = jecrc_df[jecrc_df["College"].fillna("").astype(str).str.contains("JECRC", case=False, na=False)].copy()
    if jecrc_df.empty or eval_df is None or eval_df.empty:
        return empty_batch_weekly, empty_track_weekly, empty_batch_summary

    jecrc_df["Identity_Key"] = jecrc_df.apply(
        lambda row: build_identity_key(row.get("Superset ID", ""), row.get("Candidate Name", "")),
        axis=1,
    )
    jecrc_identity_keys = {value for value in jecrc_df["Identity_Key"].tolist() if value}
    if not jecrc_identity_keys:
        return empty_batch_weekly, empty_track_weekly, empty_batch_summary

    eval_working = eval_df.copy()
    eval_working["Superset ID"] = eval_working.get("Superset ID", "").apply(normalize_superset_id) if "Superset ID" in eval_working.columns else ""
    eval_working["Candidate Name"] = eval_working.get("Candidate Name", "").fillna("").astype(str).str.strip()
    eval_working["Identity_Key"] = eval_working.apply(
        lambda row: build_identity_key(row.get("Superset ID", ""), row.get("Candidate Name", "")),
        axis=1,
    )
    eval_working = eval_working[eval_working["Identity_Key"].isin(jecrc_identity_keys)].copy()
    if eval_working.empty:
        return empty_batch_weekly, empty_track_weekly, empty_batch_summary

    eval_working["Assigned Batch"] = eval_working.get("Assigned Batch", "Unknown Batch").fillna("Unknown Batch").astype(str).str.strip()
    eval_working["AssessmentWeek"] = eval_working.get("AssessmentWeek", "").fillna("").astype(str).str.strip()
    eval_working["EvaluationTrack"] = eval_working.get("EvaluationTrack", "").apply(task2_assessment.normalize_evaluation_track)
    eval_working["Score"] = pd.to_numeric(eval_working.get("Score", np.nan), errors="coerce")
    eval_working = eval_working[eval_working["AssessmentWeek"].ne("")].copy()
    if eval_working.empty:
        return empty_batch_weekly, empty_track_weekly, empty_batch_summary

    assessment_rows = eval_working[eval_working["EvaluationTrack"].eq(task2_assessment.GITHUB_ASSESSMENT_TRACK)].copy()
    assignment_rows = eval_working[eval_working["EvaluationTrack"].eq(task2_assessment.GITHUB_ASSIGNMENT_TRACK)].copy()

    if assessment_rows.empty:
        batch_weekly_df = empty_batch_weekly
    else:
        batch_weekly_df = (
            assessment_rows.groupby(["Assigned Batch", "AssessmentWeek"], dropna=False)
            .agg(
                Average_Score=("Score", "mean"),
                Submission_Count=("Identity_Key", "nunique"),
            )
            .reset_index()
            .rename(
                columns={
                    "AssessmentWeek": "Assessment Week",
                    "Average_Score": "Average Score",
                    "Submission_Count": "Submission Count",
                }
            )
        )
        batch_weekly_df["Average Score"] = pd.to_numeric(batch_weekly_df["Average Score"], errors="coerce").round(1)
        batch_weekly_df["__batch_sort"] = batch_weekly_df["Assigned Batch"].apply(batch_sort_key)
        batch_weekly_df["__week_sort"] = batch_weekly_df["Assessment Week"].apply(week_sort_key)
        batch_weekly_df = batch_weekly_df.sort_values(by=["__batch_sort", "__week_sort"]).drop(columns=["__batch_sort", "__week_sort"]).reset_index(drop=True)

    track_weekly_source = eval_working.copy()
    if track_weekly_source.empty:
        track_weekly_df = empty_track_weekly
    else:
        track_weekly_df = (
            track_weekly_source.groupby(["EvaluationTrack", "AssessmentWeek"], dropna=False)
            .agg(
                Average_Score=("Score", "mean"),
                Submission_Count=("Identity_Key", "nunique"),
            )
            .reset_index()
            .rename(
                columns={
                    "EvaluationTrack": "Track",
                    "AssessmentWeek": "Assessment Week",
                    "Average_Score": "Average Score",
                    "Submission_Count": "Submission Count",
                }
            )
        )
        track_weekly_df["Average Score"] = pd.to_numeric(track_weekly_df["Average Score"], errors="coerce").round(1)
        track_weekly_df["Track"] = track_weekly_df["Track"].replace(
            {
                task2_assessment.GITHUB_ASSESSMENT_TRACK: "Assessment",
                task2_assessment.GITHUB_ASSIGNMENT_TRACK: "Assignment",
            }
        )
        track_weekly_df["__track_sort"] = track_weekly_df["Track"].apply(track_sort_key)
        track_weekly_df["__week_sort"] = track_weekly_df["Assessment Week"].apply(week_sort_key)
        track_weekly_df = track_weekly_df.sort_values(by=["__track_sort", "__week_sort"]).drop(columns=["__track_sort", "__week_sort"]).reset_index(drop=True)

    trainer_feedback_by_batch = (
        jecrc_df.groupby("Assigned Batch", dropna=False)
        .agg(Average_Trainer_Feedback_Score=("Trainer Feedback Score", "mean"))
        .reset_index()
    )
    trainer_feedback_by_batch["Average Trainer Feedback Score"] = pd.to_numeric(
        trainer_feedback_by_batch["Average_Trainer_Feedback_Score"],
        errors="coerce",
    ).round(1)

    assessment_count_by_batch = (
        assessment_rows.groupby("Assigned Batch", dropna=False)["Identity_Key"].nunique().reset_index(name="Assessments Done")
        if not assessment_rows.empty
        else pd.DataFrame(columns=["Assigned Batch", "Assessments Done"])
    )
    assignment_count_by_batch = (
        assignment_rows.groupby("Assigned Batch", dropna=False)["Identity_Key"].nunique().reset_index(name="Assignments Done")
        if not assignment_rows.empty
        else pd.DataFrame(columns=["Assigned Batch", "Assignments Done"])
    )
    average_weekly_assessment_by_batch = (
        batch_weekly_df.groupby("Assigned Batch", dropna=False)["Average Score"].mean().reset_index(name="Average Weekly Assessment Score")
        if not batch_weekly_df.empty
        else pd.DataFrame(columns=["Assigned Batch", "Average Weekly Assessment Score"])
    )

    batch_execution_df = jecrc_df[["Assigned Batch"]].drop_duplicates().copy()
    for frame in [assessment_count_by_batch, assignment_count_by_batch, trainer_feedback_by_batch[["Assigned Batch", "Average Trainer Feedback Score"]], average_weekly_assessment_by_batch]:
        batch_execution_df = batch_execution_df.merge(frame, on="Assigned Batch", how="left")
    for numeric_column in ["Assessments Done", "Assignments Done"]:
        batch_execution_df[numeric_column] = pd.to_numeric(batch_execution_df.get(numeric_column, 0), errors="coerce").fillna(0).astype(int)
    for numeric_column in ["Average Trainer Feedback Score", "Average Weekly Assessment Score"]:
        batch_execution_df[numeric_column] = pd.to_numeric(batch_execution_df.get(numeric_column, np.nan), errors="coerce").round(1)
    batch_execution_df["__sort_key"] = batch_execution_df["Assigned Batch"].apply(batch_sort_key)
    batch_execution_df = batch_execution_df.sort_values(by="__sort_key").drop(columns="__sort_key").reset_index(drop=True)

    return batch_weekly_df, track_weekly_df, batch_execution_df


def build_jecrc_attendance_progress_frame(filtered_df, att_data):
    empty_df = pd.DataFrame(columns=["Assigned Batch", "Attendance Week", "Average Attendance %"])
    raw_attendance_df = dataframe_from_session(att_data)
    if raw_attendance_df.empty:
        return empty_df

    jecrc_roster_df = filtered_df.copy()
    if "College" in jecrc_roster_df.columns:
        jecrc_roster_df = jecrc_roster_df[
            jecrc_roster_df["College"].fillna("").astype(str).str.contains("JECRC", case=False, na=False)
        ].copy()
    if jecrc_roster_df.empty:
        return empty_df

    attendance_name_col = find_first_column(raw_attendance_df, ["Candidate Name", "Trainee", "Student Name"])
    if attendance_name_col is None:
        return empty_df

    attendance_working = add_identity_fields(raw_attendance_df, attendance_name_col, ["Superset ID", "SupersetID", "Student ID", "Candidate ID"])
    if attendance_working.empty:
        return empty_df

    attendance_working["Identity_Key"] = attendance_working.apply(
        lambda row: build_identity_key(row.get("Superset ID", ""), row.get("Candidate Name", "")),
        axis=1,
    )
    jecrc_roster_df["Identity_Key"] = jecrc_roster_df.apply(
        lambda row: build_identity_key(row.get("Superset ID", ""), row.get("Candidate Name", "")),
        axis=1,
    )
    batch_lookup = jecrc_roster_df[["Identity_Key", "Assigned Batch"]].drop_duplicates()
    attendance_working = attendance_working.merge(batch_lookup, on="Identity_Key", how="inner")
    if attendance_working.empty:
        return empty_df

    date_cols = attendance_date_columns(attendance_working)
    if not date_cols:
        return empty_df

    def parse_attendance_date(value):
        parsed = pd.to_datetime(str(value), errors="coerce", dayfirst=True)
        return parsed

    dated_columns = [(column, parse_attendance_date(column)) for column in date_cols]
    valid_dated_columns = [item for item in dated_columns if pd.notna(item[1])]
    if valid_dated_columns:
        valid_dated_columns = sorted(valid_dated_columns, key=lambda item: item[1])
        ordered_date_cols = [item[0] for item in valid_dated_columns]
        week_lookup = {}
        distinct_weeks = []
        for _, parsed_date in valid_dated_columns:
            week_token = f"{parsed_date.isocalendar().year}-W{int(parsed_date.isocalendar().week):02d}"
            if week_token not in distinct_weeks:
                distinct_weeks.append(week_token)
        friendly_week_map = {token: f"Week {idx + 1}" for idx, token in enumerate(distinct_weeks)}
        for column_name, parsed_date in valid_dated_columns:
            week_token = f"{parsed_date.isocalendar().year}-W{int(parsed_date.isocalendar().week):02d}"
            week_lookup[column_name] = friendly_week_map[week_token]
    else:
        ordered_date_cols = list(date_cols)
        week_lookup = {column_name: f"Week {(idx // 7) + 1}" for idx, column_name in enumerate(ordered_date_cols)}

    presence_frame = attendance_working[ordered_date_cols].astype(str).apply(
        lambda column: column.str.strip().str.upper().map({"P": 1, "A": 0, "L": 0, "OD": 1})
    )
    presence_frame = presence_frame.where(presence_frame.isin([0, 1]))
    attendance_working = attendance_working.reset_index(drop=True)
    presence_frame = presence_frame.reset_index(drop=True)
    attendance_working = pd.concat([attendance_working[["Assigned Batch"]], presence_frame], axis=1)

    records = []
    for week_label in sorted(set(week_lookup.values()), key=week_sort_key):
        week_columns = [column_name for column_name, mapped_week in week_lookup.items() if mapped_week == week_label]
        if not week_columns:
            continue
        week_scores = attendance_working.groupby("Assigned Batch", dropna=False)[week_columns].mean().mean(axis=1, skipna=True).mul(100)
        for assigned_batch, value in week_scores.items():
            if pd.isna(value):
                continue
            records.append(
                {
                    "Assigned Batch": assigned_batch,
                    "Attendance Week": week_label,
                    "Average Attendance %": round(float(value), 1),
                }
            )

    if not records:
        return empty_df

    result_df = pd.DataFrame(records)
    result_df["__batch_sort"] = result_df["Assigned Batch"].apply(batch_sort_key)
    result_df["__week_sort"] = result_df["Attendance Week"].apply(week_sort_key)
    result_df = result_df.sort_values(by=["__batch_sort", "__week_sort"]).drop(columns=["__batch_sort", "__week_sort"]).reset_index(drop=True)
    return result_df


def build_college_weekly_signal_frames(filtered_df, eval_df, college_pattern, default_college_label):
    empty_batch_weekly = pd.DataFrame(columns=["College", "Assigned Batch", "Assessment Week", "Average Score", "Submission Count"])
    empty_track_weekly = pd.DataFrame(columns=["College", "Track", "Assessment Week", "Average Score", "Submission Count"])
    empty_weekly_counts = pd.DataFrame(
        columns=["College", "Assessment Week", "Assessments Done", "Assignments Done", "Trainer Feedback Provided"]
    )

    college_df = filtered_df.copy()
    if "College" in college_df.columns:
        college_df = college_df[
            college_df["College"].fillna("").astype(str).str.contains(college_pattern, case=False, na=False)
        ].copy()
    if college_df.empty or eval_df is None or eval_df.empty:
        return empty_batch_weekly, empty_track_weekly, empty_weekly_counts

    resolved_college = (
        college_df["College"].dropna().astype(str).iloc[0]
        if "College" in college_df.columns and not college_df["College"].dropna().empty
        else default_college_label
    )

    college_df["Identity_Key"] = college_df.apply(
        lambda row: build_identity_key(row.get("Superset ID", ""), row.get("Candidate Name", "")),
        axis=1,
    )
    college_identity_keys = {value for value in college_df["Identity_Key"].tolist() if value}
    if not college_identity_keys:
        return empty_batch_weekly, empty_track_weekly, empty_weekly_counts

    eval_working = eval_df.copy()
    eval_working["Superset ID"] = (
        eval_working.get("Superset ID", "").apply(normalize_superset_id)
        if "Superset ID" in eval_working.columns
        else ""
    )
    eval_working["Candidate Name"] = eval_working.get("Candidate Name", "").fillna("").astype(str).str.strip()
    eval_working["Identity_Key"] = eval_working.apply(
        lambda row: build_identity_key(row.get("Superset ID", ""), row.get("Candidate Name", "")),
        axis=1,
    )
    eval_working = eval_working[eval_working["Identity_Key"].isin(college_identity_keys)].copy()
    if eval_working.empty:
        return empty_batch_weekly, empty_track_weekly, empty_weekly_counts

    eval_working["Assigned Batch"] = (
        eval_working.get("Assigned Batch", "Unknown Batch").fillna("Unknown Batch").astype(str).str.strip()
    )
    eval_working["AssessmentWeek"] = eval_working.get("AssessmentWeek", "").fillna("").astype(str).str.strip()
    eval_working["EvaluationTrack"] = eval_working.get("EvaluationTrack", "").apply(task2_assessment.normalize_evaluation_track)
    eval_working["Score"] = pd.to_numeric(eval_working.get("Score", np.nan), errors="coerce")
    eval_working = eval_working[eval_working["AssessmentWeek"].ne("")].copy()
    if eval_working.empty:
        return empty_batch_weekly, empty_track_weekly, empty_weekly_counts

    assessment_rows = eval_working[eval_working["EvaluationTrack"].eq(task2_assessment.GITHUB_ASSESSMENT_TRACK)].copy()
    assignment_rows = eval_working[eval_working["EvaluationTrack"].eq(task2_assessment.GITHUB_ASSIGNMENT_TRACK)].copy()

    if assessment_rows.empty:
        batch_weekly_df = empty_batch_weekly
    else:
        batch_weekly_df = (
            assessment_rows.groupby(["Assigned Batch", "AssessmentWeek"], dropna=False)
            .agg(
                Average_Score=("Score", "mean"),
                Submission_Count=("Identity_Key", "nunique"),
            )
            .reset_index()
            .rename(
                columns={
                    "AssessmentWeek": "Assessment Week",
                    "Average_Score": "Average Score",
                    "Submission_Count": "Submission Count",
                }
            )
        )
        batch_weekly_df["Average Score"] = pd.to_numeric(batch_weekly_df["Average Score"], errors="coerce").round(1)
        batch_weekly_df.insert(0, "College", resolved_college)
        batch_weekly_df["__batch_sort"] = batch_weekly_df["Assigned Batch"].apply(batch_sort_key)
        batch_weekly_df["__week_sort"] = batch_weekly_df["Assessment Week"].apply(week_sort_key)
        batch_weekly_df = (
            batch_weekly_df.sort_values(by=["__batch_sort", "__week_sort"])
            .drop(columns=["__batch_sort", "__week_sort"])
            .reset_index(drop=True)
        )

    track_source = eval_working.copy()
    if track_source.empty:
        track_weekly_df = empty_track_weekly
    else:
        track_weekly_df = (
            track_source.groupby(["EvaluationTrack", "AssessmentWeek"], dropna=False)
            .agg(
                Average_Score=("Score", "mean"),
                Submission_Count=("Identity_Key", "nunique"),
            )
            .reset_index()
            .rename(
                columns={
                    "EvaluationTrack": "Track",
                    "AssessmentWeek": "Assessment Week",
                    "Average_Score": "Average Score",
                    "Submission_Count": "Submission Count",
                }
            )
        )
        track_weekly_df["Average Score"] = pd.to_numeric(track_weekly_df["Average Score"], errors="coerce").round(1)
        track_weekly_df["Track"] = track_weekly_df["Track"].replace(
            {
                task2_assessment.GITHUB_ASSESSMENT_TRACK: "Assessment",
                task2_assessment.GITHUB_ASSIGNMENT_TRACK: "Assignment",
            }
        )
        track_weekly_df.insert(0, "College", resolved_college)
        track_weekly_df["__track_sort"] = track_weekly_df["Track"].apply(track_sort_key)
        track_weekly_df["__week_sort"] = track_weekly_df["Assessment Week"].apply(week_sort_key)
        track_weekly_df = (
            track_weekly_df.sort_values(by=["__track_sort", "__week_sort"])
            .drop(columns=["__track_sort", "__week_sort"])
            .reset_index(drop=True)
        )

    feedback_identity_keys = set(
        college_df[
            pd.to_numeric(college_df.get("Trainer Feedback Score", np.nan), errors="coerce").fillna(0).gt(0)
        ]["Identity_Key"].tolist()
    )
    weekly_records = []
    ordered_weeks = sorted(eval_working["AssessmentWeek"].dropna().astype(str).unique().tolist(), key=week_sort_key)
    for week_label in ordered_weeks:
        assessment_count = 0
        if not assessment_rows.empty:
            assessment_count = int(
                assessment_rows[assessment_rows["AssessmentWeek"].eq(week_label)]["Identity_Key"].nunique()
            )
        assignment_count = 0
        if not assignment_rows.empty:
            assignment_count = int(
                assignment_rows[assignment_rows["AssessmentWeek"].eq(week_label)]["Identity_Key"].nunique()
            )
        week_identity_keys = set(
            eval_working[eval_working["AssessmentWeek"].eq(week_label)]["Identity_Key"].dropna().astype(str).tolist()
        )
        trainer_feedback_count = len(week_identity_keys & feedback_identity_keys)
        weekly_records.append(
            {
                "College": resolved_college,
                "Assessment Week": week_label,
                "Assessments Done": assessment_count,
                "Assignments Done": assignment_count,
                "Trainer Feedback Provided": trainer_feedback_count,
            }
        )
    weekly_counts_df = pd.DataFrame(weekly_records) if weekly_records else empty_weekly_counts

    return batch_weekly_df, track_weekly_df, weekly_counts_df


def build_college_attendance_progress_frame(filtered_df, att_data, college_pattern, default_college_label):
    empty_df = pd.DataFrame(columns=["College", "Assigned Batch", "Attendance Week", "Average Attendance %"])
    raw_attendance_df = dataframe_from_session(att_data)
    if raw_attendance_df.empty:
        return empty_df

    college_roster_df = filtered_df.copy()
    if "College" in college_roster_df.columns:
        college_roster_df = college_roster_df[
            college_roster_df["College"].fillna("").astype(str).str.contains(college_pattern, case=False, na=False)
        ].copy()
    if college_roster_df.empty:
        return empty_df

    resolved_college = (
        college_roster_df["College"].dropna().astype(str).iloc[0]
        if "College" in college_roster_df.columns and not college_roster_df["College"].dropna().empty
        else default_college_label
    )

    attendance_name_col = find_first_column(raw_attendance_df, ["Candidate Name", "Trainee", "Student Name"])
    if attendance_name_col is None:
        return empty_df

    attendance_working = add_identity_fields(
        raw_attendance_df,
        attendance_name_col,
        ["Superset ID", "SupersetID", "Student ID", "Candidate ID"],
    )
    if attendance_working.empty:
        return empty_df

    attendance_working["Identity_Key"] = attendance_working.apply(
        lambda row: build_identity_key(row.get("Superset ID", ""), row.get("Candidate Name", "")),
        axis=1,
    )
    college_roster_df["Identity_Key"] = college_roster_df.apply(
        lambda row: build_identity_key(row.get("Superset ID", ""), row.get("Candidate Name", "")),
        axis=1,
    )
    batch_lookup = college_roster_df[["Identity_Key", "Assigned Batch"]].drop_duplicates()
    attendance_working = attendance_working.merge(batch_lookup, on="Identity_Key", how="inner")
    if attendance_working.empty:
        return empty_df

    date_cols = attendance_date_columns(attendance_working)
    if not date_cols:
        return empty_df

    def parse_attendance_date(value):
        return pd.to_datetime(str(value), errors="coerce", dayfirst=True)

    dated_columns = [(column, parse_attendance_date(column)) for column in date_cols]
    valid_dated_columns = [item for item in dated_columns if pd.notna(item[1])]
    if valid_dated_columns:
        valid_dated_columns = sorted(valid_dated_columns, key=lambda item: item[1])
        ordered_date_cols = [item[0] for item in valid_dated_columns]
        distinct_weeks = []
        week_lookup = {}
        for _, parsed_date in valid_dated_columns:
            week_token = f"{parsed_date.isocalendar().year}-W{int(parsed_date.isocalendar().week):02d}"
            if week_token not in distinct_weeks:
                distinct_weeks.append(week_token)
        friendly_week_map = {token: f"Week {idx + 1}" for idx, token in enumerate(distinct_weeks)}
        for column_name, parsed_date in valid_dated_columns:
            week_token = f"{parsed_date.isocalendar().year}-W{int(parsed_date.isocalendar().week):02d}"
            week_lookup[column_name] = friendly_week_map[week_token]
    else:
        ordered_date_cols = list(date_cols)
        week_lookup = {column_name: f"Week {(idx // 7) + 1}" for idx, column_name in enumerate(ordered_date_cols)}

    presence_frame = attendance_working[ordered_date_cols].astype(str).apply(
        lambda column: column.str.strip().str.upper().map({"P": 1, "A": 0, "L": 0, "OD": 1})
    )
    presence_frame = presence_frame.where(presence_frame.isin([0, 1]))
    attendance_working = attendance_working.reset_index(drop=True)
    presence_frame = presence_frame.reset_index(drop=True)
    attendance_working = pd.concat([attendance_working[["Assigned Batch"]], presence_frame], axis=1)

    records = []
    for week_label in sorted(set(week_lookup.values()), key=week_sort_key):
        week_columns = [column_name for column_name, mapped_week in week_lookup.items() if mapped_week == week_label]
        if not week_columns:
            continue
        week_scores = (
            attendance_working.groupby("Assigned Batch", dropna=False)[week_columns]
            .mean()
            .mean(axis=1, skipna=True)
            .mul(100)
        )
        for assigned_batch, value in week_scores.items():
            if pd.isna(value):
                continue
            records.append(
                {
                    "College": resolved_college,
                    "Assigned Batch": assigned_batch,
                    "Attendance Week": week_label,
                    "Average Attendance %": round(float(value), 1),
                }
            )

    if not records:
        return empty_df

    result_df = pd.DataFrame(records)
    result_df["__batch_sort"] = result_df["Assigned Batch"].apply(batch_sort_key)
    result_df["__week_sort"] = result_df["Attendance Week"].apply(week_sort_key)
    result_df = (
        result_df.sort_values(by=["__batch_sort", "__week_sort"])
        .drop(columns=["__batch_sort", "__week_sort"])
        .reset_index(drop=True)
    )
    return result_df


def build_dashboard_report_bytes(
    roster_df,
    filtered_df,
    quality_metrics,
    insights,
    selected_scope,
    selected_college,
    selected_batch,
    current_quadrant_filter,
    eval_df,
    att_data,
):
    counts = filtered_df["Quadrant"].value_counts() if "Quadrant" in filtered_df.columns else pd.Series(dtype=int)
    thin_border = Border(
        left=Side(style="thin", color="D7E3F4"),
        right=Side(style="thin", color="D7E3F4"),
        top=Side(style="thin", color="D7E3F4"),
        bottom=Side(style="thin", color="D7E3F4"),
    )
    card_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="medium", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    quadrant_rows = [
        {
            "Metric": "Deployable Candidates",
            "Count": int(counts.get("Deployable Candidates", 0)),
            "Definition": QUADRANT_ACTIONS["Deployable Candidates"],
            "Quadrant": "Deployable Candidates",
            "Accent": QUADRANT_COLORS["Deployable Candidates"],
        },
        {
            "Metric": "Progressing Candidates",
            "Count": int(counts.get("Progressing Candidates", 0)),
            "Definition": QUADRANT_ACTIONS["Progressing Candidates"],
            "Quadrant": "Progressing Candidates",
            "Accent": QUADRANT_COLORS["Progressing Candidates"],
        },
        {
            "Metric": "Basic Competency",
            "Count": int(counts.get("Basic Competency", 0)),
            "Definition": QUADRANT_ACTIONS["Basic Competency"],
            "Quadrant": "Basic Competency",
            "Accent": QUADRANT_COLORS["Basic Competency"],
        },
        {
            "Metric": "Critical Intervention",
            "Count": int(counts.get("Critical Intervention", 0)),
            "Definition": QUADRANT_ACTIONS["Critical Intervention"],
            "Quadrant": "Critical Intervention",
            "Accent": QUADRANT_COLORS["Critical Intervention"],
        },
    ]
    quadrant_counts_df = pd.DataFrame([{key: row[key] for key in ["Metric", "Count", "Definition", "Quadrant"]} for row in quadrant_rows])

    summary_rows = [
        {"Metric": "Population Scope", "Value": selected_scope},
        {"Metric": "College Filter", "Value": selected_college},
        {"Metric": "Batch Filter", "Value": selected_batch},
        {"Metric": "Quadrant Filter", "Value": current_quadrant_filter},
        {"Metric": "Weighted Model", "Value": "Assessment Composite 50% | Assignment GitHub 35% | Attendance 10% | Trainer Feedback 5%"},
        {"Metric": "Quadrant Bucket Basis", "Value": "Performance Score grade band: A+/A Deployable | B Progressing | C Basic Competency | F Critical Intervention"},
        {"Metric": "Total Students", "Value": int(len(filtered_df))},
        {"Metric": "Average Assessment Composite Score", "Value": round(pd.to_numeric(filtered_df.get("Assessment Composite Score"), errors="coerce").mean(), 1)},
        {"Metric": "Average Assignment GitHub Score", "Value": round(pd.to_numeric(filtered_df.get("Assignment GitHub Score"), errors="coerce").mean(), 1)},
        {"Metric": "Average Attendance %", "Value": round(pd.to_numeric(filtered_df.get("Attendance %"), errors="coerce").mean(), 1)},
        {"Metric": "Average Trainer Feedback Score", "Value": round(pd.to_numeric(filtered_df.get("Trainer Feedback Score"), errors="coerce").mean(), 1)},
        {"Metric": "Average Performance Score", "Value": round(pd.to_numeric(filtered_df.get("Performance Score"), errors="coerce").mean(), 1)},
        {"Metric": "Average Overall Performance Score", "Value": round(pd.to_numeric(filtered_df.get("Overall Performance Score"), errors="coerce").mean(), 1)},
    ]
    summary_df = pd.DataFrame(summary_rows)

    signal_integrity_df = pd.DataFrame([{"Metric": key, "Value": value} for key, value in quality_metrics.items()])
    insights_df = pd.DataFrame([{"Headline": insight["headline"], "Detail": insight["detail"]} for insight in insights])

    plot_columns = [
        "Superset ID",
        "Candidate Name",
        "College",
        "Assigned Batch",
        "Programming Language",
        "Quadrant",
        "Performance Score",
        "Overall Performance Score",
        "Assessment Composite Score",
        "Assignment GitHub Score",
        "Attendance %",
        "Trainer Feedback Score",
    ]
    quadrant_plot_df = filtered_df[[column for column in plot_columns if column in filtered_df.columns]].copy()

    candidate_map_rows = []
    roster_lookup = filtered_df.copy()
    for row in quadrant_rows:
        members = roster_lookup[roster_lookup["Quadrant"] == row["Quadrant"]].copy()
        member_names = members["Candidate Name"].astype(str).str.strip().tolist() if "Candidate Name" in members.columns else []
        member_ids = members["Superset ID"].astype(str).str.strip().tolist() if "Superset ID" in members.columns else []
        candidate_details = []
        max_len = max(len(member_names), len(member_ids))
        for idx in range(max_len):
            superset_id = member_ids[idx] if idx < len(member_ids) else ""
            candidate_name = member_names[idx] if idx < len(member_names) else ""
            combined_detail = " - ".join(part for part in [superset_id, candidate_name] if part)
            if combined_detail:
                candidate_details.append(combined_detail)
        candidate_map_rows.append(
            {
                "Metric": row["Metric"],
                "Count": row["Count"],
                "Candidate Details": "\n".join(candidate_details),
            }
        )
    metric_candidate_map_df = pd.DataFrame(candidate_map_rows)

    working_filtered_df = filtered_df.copy()
    jecrc_batch_df = working_filtered_df.copy()
    if "College" in jecrc_batch_df.columns:
        jecrc_batch_df = jecrc_batch_df[
            jecrc_batch_df["College"].fillna("").astype(str).str.contains("JECRC", case=False, na=False)
        ].copy()
    else:
        jecrc_batch_df = pd.DataFrame()

    if not jecrc_batch_df.empty:
        jecrc_batch_summary_df = (
            jecrc_batch_df.assign(
                Assigned_Batch=jecrc_batch_df.get("Assigned Batch", "Unknown Batch").fillna("Unknown Batch").astype(str).str.strip(),
                Attendance_Numeric=pd.to_numeric(jecrc_batch_df.get("Attendance %"), errors="coerce"),
                Overall_Numeric=pd.to_numeric(jecrc_batch_df.get("Overall Performance Score"), errors="coerce"),
            )
            .groupby(["Assigned_Batch", "Quadrant"], dropna=False)
            .agg(
                Student_Count=("Superset ID", "count"),
                Average_Attendance=("Attendance_Numeric", "mean"),
                Average_Overall_Performance=("Overall_Numeric", "mean"),
            )
            .reset_index()
            .pivot_table(
                index="Assigned_Batch",
                columns="Quadrant",
                values="Student_Count",
                aggfunc="sum",
                fill_value=0,
            )
            .reset_index()
        )
        base_batch_metrics_df = (
            jecrc_batch_df.assign(
                Assigned_Batch=jecrc_batch_df.get("Assigned Batch", "Unknown Batch").fillna("Unknown Batch").astype(str).str.strip(),
                Attendance_Numeric=pd.to_numeric(jecrc_batch_df.get("Attendance %"), errors="coerce"),
                Overall_Numeric=pd.to_numeric(jecrc_batch_df.get("Overall Performance Score"), errors="coerce"),
            )
            .groupby("Assigned_Batch", dropna=False)
            .agg(
                Student_Count=("Superset ID", "count"),
                Average_Attendance=("Attendance_Numeric", "mean"),
                Average_Overall_Performance=("Overall_Numeric", "mean"),
            )
            .reset_index()
        )
        jecrc_batch_summary_df = base_batch_metrics_df.merge(
            jecrc_batch_summary_df,
            on="Assigned_Batch",
            how="left",
        ).rename(columns={"Assigned_Batch": "Assigned Batch", "Student_Count": "Student Count"})
        for bucket_name in QUADRANT_ORDER:
            if bucket_name not in jecrc_batch_summary_df.columns:
                jecrc_batch_summary_df[bucket_name] = 0
            jecrc_batch_summary_df[bucket_name] = pd.to_numeric(jecrc_batch_summary_df[bucket_name], errors="coerce").fillna(0).astype(int)
        jecrc_batch_summary_df["Average Attendance %"] = pd.to_numeric(jecrc_batch_summary_df["Average_Attendance"], errors="coerce").round(1)
        jecrc_batch_summary_df["Average Overall Performance Score"] = pd.to_numeric(jecrc_batch_summary_df["Average_Overall_Performance"], errors="coerce").round(1)
        jecrc_batch_summary_df = jecrc_batch_summary_df[
            [
                "Assigned Batch",
                "Student Count",
                "Average Attendance %",
                "Average Overall Performance Score",
                "Deployable Candidates",
                "Progressing Candidates",
                "Basic Competency",
                "Critical Intervention",
            ]
        ]
        jecrc_batch_summary_df["__sort_key"] = jecrc_batch_summary_df["Assigned Batch"].apply(batch_sort_key)
        jecrc_batch_summary_df = jecrc_batch_summary_df.sort_values(by="__sort_key").drop(columns="__sort_key").reset_index(drop=True)
    else:
        jecrc_batch_summary_df = pd.DataFrame(
            columns=[
                "Assigned Batch",
                "Student Count",
                "Average Attendance %",
                "Average Overall Performance Score",
                "Deployable Candidates",
                "Progressing Candidates",
                "Basic Competency",
                "Critical Intervention",
            ]
        )

    galgotias_batch_df = working_filtered_df.copy()
    if "College" in galgotias_batch_df.columns:
        galgotias_batch_df = galgotias_batch_df[
            galgotias_batch_df["College"].fillna("").astype(str).str.contains("Galgotias", case=False, na=False)
        ].copy()
    else:
        galgotias_batch_df = pd.DataFrame()

    if not galgotias_batch_df.empty:
        galgotias_batch_summary_df = (
            galgotias_batch_df.assign(
                Assigned_Batch=galgotias_batch_df.get("Assigned Batch", "Unknown Batch").fillna("Unknown Batch").astype(str).str.strip(),
                Attendance_Numeric=pd.to_numeric(galgotias_batch_df.get("Attendance %"), errors="coerce"),
                Overall_Numeric=pd.to_numeric(galgotias_batch_df.get("Overall Performance Score"), errors="coerce"),
            )
            .groupby("Assigned_Batch", dropna=False)
            .agg(
                Student_Count=("Superset ID", "count"),
                Average_Attendance=("Attendance_Numeric", "mean"),
                Average_Overall_Performance=("Overall_Numeric", "mean"),
            )
            .reset_index()
            .rename(columns={"Assigned_Batch": "Assigned Batch", "Student_Count": "Student Count"})
        )
        galgotias_batch_summary_df["Average Attendance %"] = galgotias_batch_summary_df["Average_Attendance"].round(1)
        galgotias_batch_summary_df["Average Overall Performance Score"] = galgotias_batch_summary_df["Average_Overall_Performance"].round(1)
        galgotias_batch_summary_df = galgotias_batch_summary_df[
            ["Assigned Batch", "Student Count", "Average Attendance %", "Average Overall Performance Score"]
        ]
        galgotias_batch_summary_df["__sort_key"] = galgotias_batch_summary_df["Assigned Batch"].apply(batch_sort_key)
        galgotias_batch_summary_df = galgotias_batch_summary_df.sort_values(by="__sort_key").drop(columns="__sort_key").reset_index(drop=True)
    else:
        galgotias_batch_summary_df = pd.DataFrame(
            columns=["Assigned Batch", "Student Count", "Average Attendance %", "Average Overall Performance Score"]
        )

    jecrc_batch_weekly_df, jecrc_track_weekly_df, jecrc_execution_summary_df = build_jecrc_weekly_bifurcation_frames(working_filtered_df, eval_df)
    jecrc_attendance_progress_df = build_jecrc_attendance_progress_frame(working_filtered_df, att_data)
    if jecrc_attendance_progress_df.empty and not jecrc_batch_summary_df.empty:
        jecrc_attendance_progress_df = jecrc_batch_summary_df[["Assigned Batch", "Average Attendance %"]].copy()
        jecrc_attendance_progress_df["Attendance Week"] = "Week 1"
        jecrc_attendance_progress_df = jecrc_attendance_progress_df[["Assigned Batch", "Attendance Week", "Average Attendance %"]]
    if jecrc_batch_weekly_df.empty and not jecrc_batch_summary_df.empty:
        jecrc_batch_weekly_df = jecrc_batch_summary_df[["Assigned Batch", "Average Overall Performance Score"]].copy()
        jecrc_batch_weekly_df["Assessment Week"] = "Week 1"
        jecrc_batch_weekly_df["Average Score"] = jecrc_batch_weekly_df["Average Overall Performance Score"]
        jecrc_batch_weekly_df["Submission Count"] = jecrc_batch_summary_df["Student Count"].astype(int)
        jecrc_batch_weekly_df = jecrc_batch_weekly_df[["Assigned Batch", "Assessment Week", "Average Score", "Submission Count"]]

    jecrc_dashboard_batch_weekly_df, jecrc_dashboard_track_weekly_df, jecrc_dashboard_weekly_counts_df = (
        build_college_weekly_signal_frames(working_filtered_df, eval_df, "JECRC", "JECRC")
    )
    galgotias_dashboard_batch_weekly_df, galgotias_dashboard_track_weekly_df, galgotias_dashboard_weekly_counts_df = (
        build_college_weekly_signal_frames(working_filtered_df, eval_df, "Galgotias", "Galgotias University")
    )
    jecrc_dashboard_attendance_df = build_college_attendance_progress_frame(working_filtered_df, att_data, "JECRC", "JECRC")
    galgotias_dashboard_attendance_df = build_college_attendance_progress_frame(
        working_filtered_df,
        att_data,
        "Galgotias",
        "Galgotias University",
    )

    if jecrc_dashboard_attendance_df.empty and not jecrc_batch_summary_df.empty:
        jecrc_dashboard_attendance_df = jecrc_batch_summary_df[["Assigned Batch", "Average Attendance %"]].copy()
        jecrc_dashboard_attendance_df.insert(0, "College", "JECRC")
        jecrc_dashboard_attendance_df["Attendance Week"] = "Week 1"
        jecrc_dashboard_attendance_df = jecrc_dashboard_attendance_df[
            ["College", "Assigned Batch", "Attendance Week", "Average Attendance %"]
        ]
    if galgotias_dashboard_attendance_df.empty and not galgotias_batch_summary_df.empty:
        galgotias_dashboard_attendance_df = galgotias_batch_summary_df[["Assigned Batch", "Average Attendance %"]].copy()
        galgotias_dashboard_attendance_df.insert(0, "College", "Galgotias University")
        galgotias_dashboard_attendance_df["Attendance Week"] = "Week 1"
        galgotias_dashboard_attendance_df = galgotias_dashboard_attendance_df[
            ["College", "Assigned Batch", "Attendance Week", "Average Attendance %"]
        ]
    if jecrc_dashboard_batch_weekly_df.empty and not jecrc_batch_summary_df.empty:
        jecrc_dashboard_batch_weekly_df = jecrc_batch_summary_df[["Assigned Batch", "Average Overall Performance Score", "Student Count"]].copy()
        jecrc_dashboard_batch_weekly_df.insert(0, "College", "JECRC")
        jecrc_dashboard_batch_weekly_df["Assessment Week"] = "Week 1"
        jecrc_dashboard_batch_weekly_df["Average Score"] = jecrc_dashboard_batch_weekly_df["Average Overall Performance Score"]
        jecrc_dashboard_batch_weekly_df["Submission Count"] = jecrc_dashboard_batch_weekly_df["Student Count"].astype(int)
        jecrc_dashboard_batch_weekly_df = jecrc_dashboard_batch_weekly_df[
            ["College", "Assigned Batch", "Assessment Week", "Average Score", "Submission Count"]
        ]
    if galgotias_dashboard_batch_weekly_df.empty and not galgotias_batch_summary_df.empty:
        galgotias_dashboard_batch_weekly_df = galgotias_batch_summary_df[
            ["Assigned Batch", "Average Overall Performance Score", "Student Count"]
        ].copy()
        galgotias_dashboard_batch_weekly_df.insert(0, "College", "Galgotias University")
        galgotias_dashboard_batch_weekly_df["Assessment Week"] = "Week 1"
        galgotias_dashboard_batch_weekly_df["Average Score"] = galgotias_dashboard_batch_weekly_df[
            "Average Overall Performance Score"
        ]
        galgotias_dashboard_batch_weekly_df["Submission Count"] = galgotias_dashboard_batch_weekly_df["Student Count"].astype(int)
        galgotias_dashboard_batch_weekly_df = galgotias_dashboard_batch_weekly_df[
            ["College", "Assigned Batch", "Assessment Week", "Average Score", "Submission Count"]
        ]

    dashboard_batch_weekly_df = pd.concat(
        [jecrc_dashboard_batch_weekly_df, galgotias_dashboard_batch_weekly_df],
        ignore_index=True,
    )
    dashboard_track_weekly_df = pd.concat(
        [jecrc_dashboard_track_weekly_df, galgotias_dashboard_track_weekly_df],
        ignore_index=True,
    )
    dashboard_weekly_counts_df = pd.concat(
        [jecrc_dashboard_weekly_counts_df, galgotias_dashboard_weekly_counts_df],
        ignore_index=True,
    )
    dashboard_attendance_progress_df = pd.concat(
        [jecrc_dashboard_attendance_df, galgotias_dashboard_attendance_df],
        ignore_index=True,
    )

    def apply_table_format(worksheet):
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.fill = PatternFill("solid", fgColor="0F172A")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
        for col_idx, column_cells in enumerate(worksheet.columns, start=1):
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 14), 42)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        workbook = writer.book

        summary_sheet = workbook.create_sheet("dashboard_summary", 0)
        summary_sheet.sheet_view.showGridLines = False
        summary_sheet.freeze_panes = "A6"
        summary_sheet.column_dimensions["A"].width = 18
        summary_sheet.column_dimensions["B"].width = 18
        summary_sheet.column_dimensions["C"].width = 18
        summary_sheet.column_dimensions["D"].width = 18
        summary_sheet.column_dimensions["E"].width = 18
        summary_sheet.column_dimensions["F"].width = 18
        summary_sheet.column_dimensions["G"].width = 18
        summary_sheet.column_dimensions["H"].width = 18
        summary_sheet.column_dimensions["I"].width = 4
        summary_sheet.column_dimensions["J"].width = 18
        summary_sheet.column_dimensions["K"].width = 10
        summary_sheet.column_dimensions["L"].width = 12

        for row in range(1, 80):
            fill_color = "F8FBFF" if row > 4 else "0F172A"
            for col in range(1, 9):
                summary_sheet.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill_color)

        summary_sheet.merge_cells("A1:H3")
        header_cell = summary_sheet["A1"]
        header_cell.value = "Performance Magic Quadrant Plus Dashboard Report"
        header_cell.font = Font(color="FFFFFF", bold=True, size=22)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

        summary_sheet.merge_cells("A4:H4")
        subtitle_cell = summary_sheet["A4"]
        subtitle_cell.value = "Executive intelligence export using the approved 50/35/10/5 weighted model and performance-score grade buckets."
        subtitle_cell.font = Font(color="E2E8F0", italic=True, size=11)
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

        context_rows = [
            ("Population Scope", selected_scope),
            ("College Filter", selected_college),
            ("Batch Filter", selected_batch),
            ("Quadrant Filter", current_quadrant_filter),
        ]
        summary_sheet.merge_cells("A6:D6")
        summary_sheet["A6"] = "Current Filter Context"
        summary_sheet["A6"].font = Font(bold=True, size=13, color="0F172A")
        summary_sheet["A6"].alignment = Alignment(horizontal="left", vertical="center")
        row_cursor = 7
        for label, value in context_rows:
            summary_sheet.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=2)
            summary_sheet.merge_cells(start_row=row_cursor, start_column=3, end_row=row_cursor, end_column=4)
            label_cell = summary_sheet.cell(row=row_cursor, column=1)
            value_cell = summary_sheet.cell(row=row_cursor, column=3)
            label_cell.value = label
            value_cell.value = value
            label_cell.font = Font(bold=True, color="334155")
            value_cell.font = Font(color="0F172A")
            for col in range(1, 5):
                summary_sheet.cell(row=row_cursor, column=col).fill = PatternFill("solid", fgColor="EAF2FF")
                summary_sheet.cell(row=row_cursor, column=col).border = thin_border
                summary_sheet.cell(row=row_cursor, column=col).alignment = Alignment(horizontal="left", vertical="center")
            row_cursor += 1

        summary_sheet.merge_cells("E6:H6")
        summary_sheet["E6"] = "Weighted Model"
        summary_sheet["E6"].font = Font(bold=True, size=13, color="0F172A")
        summary_sheet["E6"].alignment = Alignment(horizontal="left", vertical="center")
        summary_sheet.merge_cells("E7:H9")
        model_cell = summary_sheet["E7"]
        model_cell.value = (
            "Overall weighted model: Assessment Composite 50% | Assignment GitHub 35% | "
            "Attendance 10% | Trainer Feedback 5%\n"
            "Quadrant bucket basis: Performance Score grade band only."
        )
        model_cell.font = Font(color="0F172A", bold=True, size=12)
        model_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for row in range(7, 10):
            for col in range(5, 9):
                summary_sheet.cell(row=row, column=col).fill = PatternFill("solid", fgColor="EAF2FF")
                summary_sheet.cell(row=row, column=col).border = thin_border

        card_ranges = [
            ("A11:B16", quadrant_rows[0], "A and A+ performance band students ready for deployment conversations."),
            ("C11:D16", quadrant_rows[1], "B-grade students showing good progress and improving technical readiness."),
            ("E11:F16", quadrant_rows[2], "C-grade students with baseline competency who still need guided uplift."),
            ("G11:H16", quadrant_rows[3], "F-grade students who need immediate performance intervention."),
        ]
        for cell_range, row_meta, description in card_ranges:
            start_cell, end_cell = cell_range.split(":")
            start_col = summary_sheet[start_cell].column
            start_row = summary_sheet[start_cell].row
            end_col = summary_sheet[end_cell].column
            end_row = summary_sheet[end_cell].row
            summary_sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row + 1, end_column=end_col)
            summary_sheet.merge_cells(start_row=start_row + 2, start_column=start_col, end_row=start_row + 3, end_column=end_col)
            summary_sheet.merge_cells(start_row=start_row + 4, start_column=start_col, end_row=end_row, end_column=end_col)
            title_cell = summary_sheet.cell(row=start_row, column=start_col)
            value_cell = summary_sheet.cell(row=start_row + 2, column=start_col)
            desc_cell = summary_sheet.cell(row=start_row + 4, column=start_col)
            title_cell.value = row_meta["Metric"].upper()
            value_cell.value = row_meta["Count"]
            desc_cell.value = description
            title_cell.font = Font(bold=True, size=11, color=row_meta["Accent"].replace("#", "").upper())
            value_cell.font = Font(bold=True, size=28, color="0F172A")
            desc_cell.font = Font(size=10, color="475569")
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = summary_sheet.cell(row=row, column=col)
                    cell.fill = PatternFill("solid", fgColor="FFFFFF")
                    cell.border = card_border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            for col in range(start_col, end_col + 1):
                summary_sheet.cell(row=start_row, column=col).border = Border(
                    left=card_border.left,
                    right=card_border.right,
                    top=Side(style="medium", color=row_meta["Accent"].replace("#", "").upper()),
                    bottom=card_border.bottom,
                )

        summary_sheet.merge_cells("A18:D18")
        summary_sheet["A18"] = "Executive AI Insight Panel"
        summary_sheet["A18"].font = Font(bold=True, size=13, color="0F172A")
        insight_start_row = 19
        for insight in insights[:4]:
            summary_sheet.merge_cells(start_row=insight_start_row, start_column=1, end_row=insight_start_row, end_column=4)
            summary_sheet.merge_cells(start_row=insight_start_row + 1, start_column=1, end_row=insight_start_row + 2, end_column=4)
            headline_cell = summary_sheet.cell(row=insight_start_row, column=1)
            detail_cell = summary_sheet.cell(row=insight_start_row + 1, column=1)
            headline_cell.value = insight["headline"]
            detail_cell.value = insight["detail"]
            headline_cell.font = Font(bold=True, size=11, color="0F172A")
            detail_cell.font = Font(size=10, color="475569")
            for row in range(insight_start_row, insight_start_row + 3):
                for col in range(1, 5):
                    summary_sheet.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFFFFF")
                    summary_sheet.cell(row=row, column=col).border = thin_border
                    summary_sheet.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            insight_start_row += 4

        summary_sheet.merge_cells("E18:H18")
        summary_sheet["E18"] = "Signal Integrity Stack"
        summary_sheet["E18"].font = Font(bold=True, size=13, color="0F172A")
        signal_rows = [
            ("Population", quality_metrics.get("Population", 0)),
            ("Feedback ID Match", quality_metrics.get("Feedback ID Matches", 0)),
            ("Top Brains ID Match", quality_metrics.get("Top Brains ID Matches", 0)),
            ("Assess GitHub ID Match", quality_metrics.get("Assessment GitHub ID Matches", 0)),
            ("Assign GitHub ID Match", quality_metrics.get("Assignment GitHub ID Matches", 0)),
            ("Name Fallbacks", quality_metrics.get("Name Fallbacks", 0)),
        ]
        signal_row = 19
        for label, value in signal_rows:
            summary_sheet.merge_cells(start_row=signal_row, start_column=5, end_row=signal_row, end_column=6)
            summary_sheet.merge_cells(start_row=signal_row, start_column=7, end_row=signal_row, end_column=8)
            summary_sheet.cell(row=signal_row, column=5).value = label
            summary_sheet.cell(row=signal_row, column=7).value = value
            summary_sheet.cell(row=signal_row, column=5).font = Font(bold=True, color="334155")
            summary_sheet.cell(row=signal_row, column=7).font = Font(bold=True, color="0F172A")
            for col in range(5, 9):
                summary_sheet.cell(row=signal_row, column=col).fill = PatternFill("solid", fgColor="FFFFFF")
                summary_sheet.cell(row=signal_row, column=col).border = thin_border
                summary_sheet.cell(row=signal_row, column=col).alignment = Alignment(horizontal="left", vertical="center")
            signal_row += 2

        summary_sheet.merge_cells("A36:H36")
        summary_sheet["A36"] = "Metric-to-Candidate Mapping"
        summary_sheet["A36"].font = Font(bold=True, size=13, color="0F172A")
        summary_sheet["A37"] = "Metric"
        summary_sheet["B37"] = "Count"
        summary_sheet.merge_cells("C37:H37")
        summary_sheet["C37"] = "Candidate Details in Current View"
        for col in range(1, 9):
            summary_sheet.cell(row=37, column=col).fill = PatternFill("solid", fgColor="0F172A")
            summary_sheet.cell(row=37, column=col).font = Font(color="FFFFFF", bold=True)
            summary_sheet.cell(row=37, column=col).alignment = Alignment(horizontal="center", vertical="center")
            summary_sheet.cell(row=37, column=col).border = thin_border

        map_row = 38
        for row in candidate_map_rows:
            summary_sheet.cell(row=map_row, column=1).value = row["Metric"]
            summary_sheet.cell(row=map_row, column=2).value = row["Count"]
            summary_sheet.merge_cells(start_row=map_row, start_column=3, end_row=map_row + 2, end_column=8)
            summary_sheet.cell(row=map_row, column=3).value = row["Candidate Details"] or "No candidates in this segment."
            summary_sheet.cell(row=map_row, column=1).font = Font(bold=True, color=row["Metric"] and "0F172A")
            for row_idx in range(map_row, map_row + 3):
                for col in range(1, 9):
                    summary_sheet.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FFFFFF")
                    summary_sheet.cell(row=row_idx, column=col).border = thin_border
                    summary_sheet.cell(row=row_idx, column=col).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            summary_sheet.row_dimensions[map_row].height = 26
            summary_sheet.row_dimensions[map_row + 1].height = 48
            summary_sheet.row_dimensions[map_row + 2].height = 48
            map_row += 4

        chart_start_row = 11
        summary_sheet["J1"] = "Performance Bucket"
        summary_sheet["K1"] = "Count"
        for idx, row in enumerate(quadrant_rows, start=2):
            summary_sheet[f"J{idx}"] = row["Metric"]
            summary_sheet[f"K{idx}"] = row["Count"]
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Performance Bucket Split"
        chart.y_axis.title = "Bucket"
        chart.x_axis.title = "Candidate Count"
        chart.legend = None
        data = Reference(summary_sheet, min_col=11, min_row=1, max_row=5)
        categories = Reference(summary_sheet, min_col=10, min_row=2, max_row=5)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8.2
        chart.width = 11
        chart.varyColors = True
        summary_sheet.add_chart(chart, "E27")
        summary_sheet.column_dimensions["J"].hidden = True
        summary_sheet.column_dimensions["K"].hidden = True
        summary_sheet.column_dimensions["L"].hidden = True

        summary_df.to_excel(writer, index=False, sheet_name="summary_data")
        quadrant_counts_df.to_excel(writer, index=False, sheet_name="quadrant_counts")
        signal_integrity_df.to_excel(writer, index=False, sheet_name="signal_integrity")
        insights_df.to_excel(writer, index=False, sheet_name="executive_insights")
        metric_candidate_map_df.to_excel(writer, index=False, sheet_name="metric_candidate_map")
        quadrant_plot_df.to_excel(writer, index=False, sheet_name="quadrant_plot_data")
        roster_df.to_excel(writer, index=False, sheet_name="executive_roster")
        jecrc_batch_summary_df.to_excel(writer, index=False, sheet_name="jecrc_batch_performance")
        galgotias_batch_summary_df.to_excel(writer, index=False, sheet_name="galgotias_batch_performance")
        dashboard_batch_weekly_df.to_excel(writer, index=False, sheet_name="college_batch_weekly_performance")
        dashboard_track_weekly_df.to_excel(writer, index=False, sheet_name="college_track_weekly_performance")
        dashboard_weekly_counts_df.to_excel(writer, index=False, sheet_name="college_weekly_signal_counts")
        dashboard_attendance_progress_df.to_excel(writer, index=False, sheet_name="college_batch_weekly_attendance")

        workbook["dashboard_summary"].sheet_properties.tabColor = "1D4ED8"
        workbook["summary_data"].sheet_properties.tabColor = "2563EB"
        workbook["quadrant_counts"].sheet_properties.tabColor = "10B981"
        workbook["signal_integrity"].sheet_properties.tabColor = "7C3AED"
        workbook["executive_insights"].sheet_properties.tabColor = "F59E0B"
        workbook["metric_candidate_map"].sheet_properties.tabColor = "0EA5E9"
        workbook["quadrant_plot_data"].sheet_properties.tabColor = "64748B"
        workbook["executive_roster"].sheet_properties.tabColor = "0F172A"
        workbook["jecrc_batch_performance"].sheet_properties.tabColor = "DC2626"
        workbook["galgotias_batch_performance"].sheet_properties.tabColor = "2563EB"
        workbook["college_batch_weekly_performance"].sheet_properties.tabColor = "0EA5E9"
        workbook["college_track_weekly_performance"].sheet_properties.tabColor = "7C3AED"
        workbook["college_weekly_signal_counts"].sheet_properties.tabColor = "F59E0B"
        workbook["college_batch_weekly_attendance"].sheet_properties.tabColor = "10B981"

        for sheet_name in [
            "summary_data",
            "quadrant_counts",
            "signal_integrity",
            "executive_insights",
            "metric_candidate_map",
            "quadrant_plot_data",
            "executive_roster",
            "jecrc_batch_performance",
            "galgotias_batch_performance",
            "college_batch_weekly_performance",
            "college_track_weekly_performance",
            "college_weekly_signal_counts",
            "college_batch_weekly_attendance",
        ]:
            apply_table_format(workbook[sheet_name])

        batch_sheet = workbook["jecrc_batch_performance"]
        batch_sheet.sheet_view.showGridLines = False
        batch_sheet.freeze_panes = "A2"
        if batch_sheet.max_row <= 1:
            batch_sheet.merge_cells("A3:H5")
            empty_cell = batch_sheet["A3"]
            empty_cell.value = "No JECRC batch-level data is available in the current PMQ view."
            empty_cell.font = Font(size=12, bold=True, color="475569")
            empty_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            chart_data_end_row = batch_sheet.max_row

            attendance_chart = BarChart()
            attendance_chart.type = "col"
            attendance_chart.style = 10
            attendance_chart.title = "JECRC Batch-wise Average Attendance %"
            attendance_chart.y_axis.title = "Average Attendance Percentage"
            attendance_chart.x_axis.title = "Batch"
            attendance_chart.height = 8.5
            attendance_chart.width = 14
            attendance_chart.varyColors = False
            attendance_data = Reference(batch_sheet, min_col=3, min_row=1, max_row=chart_data_end_row)
            attendance_categories = Reference(batch_sheet, min_col=1, min_row=2, max_row=chart_data_end_row)
            attendance_chart.add_data(attendance_data, titles_from_data=True)
            attendance_chart.set_categories(attendance_categories)
            attendance_chart.legend = None
            attendance_chart.dLbls = DataLabelList()
            attendance_chart.dLbls.showVal = True
            attendance_chart.y_axis.scaling.min = 0
            attendance_chart.y_axis.scaling.max = 100
            if attendance_chart.ser:
                attendance_chart.ser[0].dPt = []
                for idx in range(chart_data_end_row - 1):
                    point = DataPoint(idx=idx)
                    point.graphicalProperties = GraphicalProperties(solidFill=JECRC_BATCH_COLORS[idx % len(JECRC_BATCH_COLORS)])
                    attendance_chart.ser[0].dPt.append(point)
            batch_sheet.add_chart(attendance_chart, "K2")

            bucket_chart = BarChart()
            bucket_chart.type = "bar"
            bucket_chart.style = 12
            bucket_chart.grouping = "stacked"
            bucket_chart.overlap = 100
            bucket_chart.title = "JECRC Batch-wise Performance Bucket Mix"
            bucket_chart.y_axis.title = "Batch"
            bucket_chart.x_axis.title = "Student Count"
            bucket_chart.height = 9.5
            bucket_chart.width = 15
            bucket_data = Reference(batch_sheet, min_col=5, min_row=1, max_col=8, max_row=chart_data_end_row)
            bucket_categories = Reference(batch_sheet, min_col=1, min_row=2, max_row=chart_data_end_row)
            bucket_chart.add_data(bucket_data, titles_from_data=True)
            bucket_chart.set_categories(bucket_categories)
            bucket_chart.legend.position = "b"
            bucket_palette = ["10B981", "7C3AED", "0EA5E9", "EF4444"]
            for idx, series in enumerate(bucket_chart.ser):
                series.graphicalProperties = GraphicalProperties(solidFill=bucket_palette[idx % len(bucket_palette)])
            batch_sheet.add_chart(bucket_chart, "K20")

            if not jecrc_attendance_progress_df.empty:
                attendance_progress_start_col = 20
                attendance_progress_start_row = 2
                attendance_pivot = jecrc_attendance_progress_df.pivot(
                    index="Attendance Week",
                    columns="Assigned Batch",
                    values="Average Attendance %",
                ).reset_index()
                for row_offset, row in enumerate([attendance_pivot.columns.tolist()] + attendance_pivot.fillna(0).values.tolist()):
                    for col_offset, value in enumerate(row):
                        batch_sheet.cell(row=attendance_progress_start_row + row_offset, column=attendance_progress_start_col + col_offset).value = value
                attendance_progress_chart = LineChart()
                attendance_progress_chart.title = "JECRC Week-on-Week Attendance Progress"
                attendance_progress_chart.y_axis.title = "Average Attendance Percentage"
                attendance_progress_chart.x_axis.title = "Week"
                attendance_progress_chart.height = 7.5
                attendance_progress_chart.width = 14
                attendance_progress_chart.y_axis.scaling.min = 0
                attendance_progress_chart.y_axis.scaling.max = 100
                attendance_progress_data = Reference(
                    batch_sheet,
                    min_col=attendance_progress_start_col + 1,
                    min_row=attendance_progress_start_row,
                    max_col=attendance_progress_start_col + len(attendance_pivot.columns) - 1,
                    max_row=attendance_progress_start_row + len(attendance_pivot),
                )
                attendance_progress_categories = Reference(
                    batch_sheet,
                    min_col=attendance_progress_start_col,
                    min_row=attendance_progress_start_row + 1,
                    max_row=attendance_progress_start_row + len(attendance_pivot),
                )
                attendance_progress_chart.add_data(attendance_progress_data, titles_from_data=True)
                attendance_progress_chart.set_categories(attendance_progress_categories)
                attendance_progress_chart.legend.position = "b"
                batch_sheet.add_chart(attendance_progress_chart, "K38")

            if not jecrc_batch_weekly_df.empty:
                performance_progress_start_col = 20
                performance_progress_start_row = 24
                performance_pivot = jecrc_batch_weekly_df.pivot(
                    index="Assessment Week",
                    columns="Assigned Batch",
                    values="Average Score",
                ).reset_index()
                for row_offset, row in enumerate([performance_pivot.columns.tolist()] + performance_pivot.fillna(0).values.tolist()):
                    for col_offset, value in enumerate(row):
                        batch_sheet.cell(row=performance_progress_start_row + row_offset, column=performance_progress_start_col + col_offset).value = value
                performance_progress_chart = LineChart()
                performance_progress_chart.title = "JECRC Week-on-Week Performance Progress"
                performance_progress_chart.y_axis.title = "Average Score"
                performance_progress_chart.x_axis.title = "Week"
                performance_progress_chart.height = 7.5
                performance_progress_chart.width = 14
                performance_progress_chart.y_axis.scaling.min = 0
                performance_progress_chart.y_axis.scaling.max = 100
                performance_progress_data = Reference(
                    batch_sheet,
                    min_col=performance_progress_start_col + 1,
                    min_row=performance_progress_start_row,
                    max_col=performance_progress_start_col + len(performance_pivot.columns) - 1,
                    max_row=performance_progress_start_row + len(performance_pivot),
                )
                performance_progress_categories = Reference(
                    batch_sheet,
                    min_col=performance_progress_start_col,
                    min_row=performance_progress_start_row + 1,
                    max_row=performance_progress_start_row + len(performance_pivot),
                )
                performance_progress_chart.add_data(performance_progress_data, titles_from_data=True)
                performance_progress_chart.set_categories(performance_progress_categories)
                performance_progress_chart.legend.position = "b"
                batch_sheet.add_chart(performance_progress_chart, "K56")
            for hidden_col in range(20, 31):
                batch_sheet.column_dimensions[get_column_letter(hidden_col)].hidden = True

        galgotias_sheet = workbook["galgotias_batch_performance"]
        galgotias_sheet.sheet_view.showGridLines = False
        galgotias_sheet.freeze_panes = "A2"
        if galgotias_sheet.max_row <= 1:
            galgotias_sheet.merge_cells("A3:D5")
            empty_cell = galgotias_sheet["A3"]
            empty_cell.value = "No Galgotias batch-level data is available in the current PMQ view."
            empty_cell.font = Font(size=12, bold=True, color="475569")
            empty_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            chart_data_end_row = galgotias_sheet.max_row

            attendance_chart = BarChart()
            attendance_chart.type = "col"
            attendance_chart.style = 10
            attendance_chart.title = "Galgotias Batch-wise Average Attendance %"
            attendance_chart.y_axis.title = "Attendance %"
            attendance_chart.x_axis.title = "Galgotias Batches"
            attendance_chart.height = 7.5
            attendance_chart.width = 11
            attendance_chart.varyColors = False
            attendance_data = Reference(galgotias_sheet, min_col=3, min_row=1, max_row=chart_data_end_row)
            attendance_categories = Reference(galgotias_sheet, min_col=1, min_row=2, max_row=chart_data_end_row)
            attendance_chart.add_data(attendance_data, titles_from_data=True)
            attendance_chart.set_categories(attendance_categories)
            attendance_chart.y_axis.scaling.min = 0
            attendance_chart.y_axis.scaling.max = 100
            if attendance_chart.ser:
                attendance_chart.ser[0].dPt = []
                for idx in range(chart_data_end_row - 1):
                    point = DataPoint(idx=idx)
                    point.graphicalProperties = GraphicalProperties(solidFill=JECRC_BATCH_COLORS[idx % len(JECRC_BATCH_COLORS)])
                    attendance_chart.ser[0].dPt.append(point)
            galgotias_sheet.add_chart(attendance_chart, "F2")

            overall_chart = BarChart()
            overall_chart.type = "col"
            overall_chart.style = 11
            overall_chart.title = "Galgotias Batch-wise Average Overall Performance"
            overall_chart.y_axis.title = "Overall Performance Score"
            overall_chart.x_axis.title = "Galgotias Batches"
            overall_chart.height = 7.5
            overall_chart.width = 11
            overall_chart.varyColors = False
            overall_data = Reference(galgotias_sheet, min_col=4, min_row=1, max_row=chart_data_end_row)
            overall_categories = Reference(galgotias_sheet, min_col=1, min_row=2, max_row=chart_data_end_row)
            overall_chart.add_data(overall_data, titles_from_data=True)
            overall_chart.set_categories(overall_categories)
            overall_chart.y_axis.scaling.min = 0
            overall_chart.y_axis.scaling.max = 100
            if overall_chart.ser:
                overall_chart.ser[0].dPt = []
                for idx in range(chart_data_end_row - 1):
                    point = DataPoint(idx=idx)
                    point.graphicalProperties = GraphicalProperties(solidFill=JECRC_BATCH_COLORS[idx % len(JECRC_BATCH_COLORS)])
                    overall_chart.ser[0].dPt.append(point)
            galgotias_sheet.add_chart(overall_chart, "F18")

    buffer.seek(0)
    return buffer.getvalue()


def render_kpi_card(title, count, description, accent):
    st.markdown(
        f"""
        <div style="background:#ffffff;border:1px solid #dbe7f5;border-top:4px solid {accent};border-radius:18px;padding:18px 18px 16px;box-shadow:0 12px 28px rgba(15,23,42,0.06);height:100%;">
            <div style="font-size:12px;font-weight:800;letter-spacing:0.08em;text-transform:uppercase;color:{accent};margin-bottom:10px;">{title}</div>
            <div style="font-size:34px;font-weight:900;color:#0f172a;line-height:1;">{count}</div>
            <div style="font-size:13px;color:#475569;margin-top:10px;line-height:1.5;">{description}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def run():
    if "task5_quadrant_filter" not in st.session_state:
        st.session_state["task5_quadrant_filter"] = "All Quadrants"
    if "task5_population_scope" not in st.session_state:
        st.session_state["task5_population_scope"] = "Active Only"

    st.markdown(
        """
        <style>
        .gm-shell {
            background: linear-gradient(135deg, #0f172a 0%, #1d4ed8 100%);
            border-radius: 22px;
            padding: 28px 32px;
            margin-bottom: 18px;
            box-shadow: 0 20px 40px rgba(15, 23, 42, 0.14);
        }
        .gm-title {
            color: #ffffff;
            font-size: 32px;
            font-weight: 900;
            margin: 0 0 8px 0;
        }
        .gm-subtitle {
            color: rgba(255,255,255,0.88);
            font-size: 15px;
            line-height: 1.6;
            max-width: 1050px;
        }
        .gm-panel {
            background: #ffffff;
            border: 1px solid #dbe7f5;
            border-radius: 18px;
            padding: 18px 20px;
            box-shadow: 0 12px 28px rgba(15,23,42,0.05);
            margin-bottom: 16px;
        }
        .gm-panel-title {
            color: #0f172a;
            font-size: 20px;
            font-weight: 800;
            margin-bottom: 6px;
        }
        .gm-panel-subtitle {
            color: #64748b;
            font-size: 14px;
            line-height: 1.6;
        }
        .gm-insight {
            background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%);
            border: 1px solid #dbe7f5;
            border-radius: 16px;
            padding: 14px 16px;
            margin-bottom: 10px;
        }
        .gm-insight-headline {
            color: #0f172a;
            font-size: 15px;
            font-weight: 800;
            margin-bottom: 4px;
        }
        .gm-insight-detail {
            color: #475569;
            font-size: 13px;
            line-height: 1.55;
        }
        </style>
        <div class="gm-shell">
            <div class="gm-title">Performance Magic Quadrant Plus</div>
            <div class="gm-subtitle">
                Executive intelligence board using Superset ID as the primary merge key and the approved weighted model:
                assessment GitHub plus Top Brains composite (50%), assignment GitHub (35%), attendance (10%), and trainer feedback (5%).
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    tracker_data = st.session_state.get("tracker_data")
    att_data = st.session_state.get("att_data")
    if tracker_data is None or att_data is None:
        st.warning("Run Task-2 and Task-3 first so this module can consume the current session datasets. Task-1 code data is pulled automatically when available.")
        return

    eval_df, eval_notices = build_combined_eval_frame()
    raw_tracker_df = dataframe_from_session(tracker_data)
    feedback_df = build_feedback_frame(raw_tracker_df)
    top_brains_df = build_top_brains_frame(raw_tracker_df)
    attendance_df = build_attendance_frame(att_data)
    assessment_github_df = build_github_track_frame(
        eval_df,
        task2_assessment.GITHUB_ASSESSMENT_TRACK,
        "Assessment GitHub Score",
        "Assessment GitHub Weeks",
        "Assessment GitHub Candidate Name",
        "Assessment GitHub Assigned Batch",
        "Assessment GitHub Technology",
    )
    assignment_github_df = build_github_track_frame(
        eval_df,
        task2_assessment.GITHUB_ASSIGNMENT_TRACK,
        "Assignment GitHub Score",
        "Assignment GitHub Weeks",
        "Assignment GitHub Candidate Name",
        "Assignment GitHub Assigned Batch",
        "Assignment GitHub Technology",
    )
    dropout_override_df = build_dropout_override_frame(eval_df)

    if attendance_df.empty:
        st.warning("The executive intelligence layer could not build a valid attendance base from the current session-state records.")
        return

    population_options = ["Active Only", "All Students"]
    if st.session_state.get("task5_population_scope") not in population_options:
        st.session_state["task5_population_scope"] = "Active Only"
    selected_scope = st.selectbox("Population Scope", population_options, key="task5_population_scope")
    merged_df = build_executive_frame(
        assessment_github_df,
        assignment_github_df,
        feedback_df,
        top_brains_df,
        attendance_df,
        dropout_override_df,
        selected_scope,
    )
    if merged_df.empty:
        st.warning("No candidates remain in the selected population scope.")
        return

    merged_df["Quadrant"] = merged_df.apply(assign_quadrant, axis=1)
    merged_df["Bubble Size"] = pd.to_numeric(merged_df["Overall Performance Score"], errors="coerce").fillna(0).apply(lambda value: max((float(value) / 4.0), 10.0))

    filter_panel = st.container(border=True)
    with filter_panel:
        st.markdown("#### Executive Filters")
        st.caption("Population scope, college, and batch filters update the weighted quadrant, data-quality view, and export roster together.")
        filter_col1, filter_col2, filter_col3 = st.columns(3)
        with filter_col1:
            st.caption("Superset-ID-first merge is active in this view.")
            college_options = ["All Colleges"] + sorted(merged_df["College"].dropna().astype(str).unique().tolist())
        if st.session_state.get("task5_college") not in college_options:
            st.session_state["task5_college"] = "All Colleges"
        with filter_col2:
            selected_college = st.selectbox("College", college_options, key="task5_college")
        scoped_df = merged_df if selected_college == "All Colleges" else merged_df[merged_df["College"] == selected_college].copy()
        batch_options = ["All Batches"] + sorted(scoped_df["Assigned Batch"].dropna().astype(str).unique().tolist())
        if st.session_state.get("task5_batch") not in batch_options:
            st.session_state["task5_batch"] = "All Batches"
        with filter_col3:
            batch_options = ["All Batches"] + sorted(scoped_df["Assigned Batch"].dropna().astype(str).unique().tolist())
            selected_batch = st.selectbox("Batch", batch_options, key="task5_batch")

    filtered_df = scoped_df.copy()
    if selected_batch != "All Batches":
        filtered_df = filtered_df[filtered_df["Assigned Batch"] == selected_batch].copy()

    if filtered_df.empty:
        st.warning("No candidates remain after applying the selected executive filters.")
        return

    if eval_notices:
        st.caption(f"GitHub normalization notices: {len(eval_notices)}. Other college uploads continue to work here as long as their Task-2 and Task-3 files are loaded in the same session.")

    quality_metrics = build_data_quality_metrics(filtered_df)
    counts = filtered_df["Quadrant"].value_counts()
    kpi_cols = st.columns(4)
    with kpi_cols[0]:
        render_kpi_card(
            "Deployable Candidates",
            counts.get("Deployable Candidates", 0),
            "A and A+ performance-grade students ready for deployment discussions.",
            QUADRANT_COLORS["Deployable Candidates"],
        )
        if st.button("View Deployable", key="task5_view_q1", use_container_width=True):
            st.session_state["task5_quadrant_filter"] = "Deployable Candidates"
            st.rerun()
    with kpi_cols[1]:
        render_kpi_card(
            "Progressing Candidates",
            counts.get("Progressing Candidates", 0),
            "B-grade performance students who are building strong momentum.",
            QUADRANT_COLORS["Progressing Candidates"],
        )
        if st.button("View Progressing", key="task5_view_q2", use_container_width=True):
            st.session_state["task5_quadrant_filter"] = "Progressing Candidates"
            st.rerun()
    with kpi_cols[2]:
        render_kpi_card(
            "Basic Competency",
            counts.get("Basic Competency", 0),
            "C-grade performance students who have cleared the baseline but still need uplift.",
            QUADRANT_COLORS["Basic Competency"],
        )
        if st.button("View Basic Competency", key="task5_view_q3", use_container_width=True):
            st.session_state["task5_quadrant_filter"] = "Basic Competency"
            st.rerun()
    with kpi_cols[3]:
        render_kpi_card(
            "Critical Intervention",
            counts.get("Critical Intervention", 0),
            "F-grade performance students requiring immediate management attention.",
            QUADRANT_COLORS["Critical Intervention"],
        )
        if st.button("View Critical", key="task5_view_q4", use_container_width=True):
            st.session_state["task5_quadrant_filter"] = "Critical Intervention"
            st.rerun()

    current_quadrant_filter = st.session_state.get("task5_quadrant_filter", "All Quadrants")
    if current_quadrant_filter not in {"All Quadrants"} | set(filtered_df["Quadrant"].unique().tolist()):
        current_quadrant_filter = "All Quadrants"
        st.session_state["task5_quadrant_filter"] = current_quadrant_filter

    reset_cols = st.columns([1, 3])
    with reset_cols[0]:
        if st.button("View All Cohort", key="task5_view_all", use_container_width=True):
            st.session_state["task5_quadrant_filter"] = "All Quadrants"
            st.rerun()

    insight_col, signal_col = st.columns([1.35, 1])
    current_insights = build_exec_insights(filtered_df, quality_metrics, selected_scope)

    with insight_col:
        st.markdown(
            """
            <div class="gm-panel">
                <div class="gm-panel-title">Executive AI Insight Panel</div>
                <div class="gm-panel-subtitle">Rule-based leadership intelligence summarizing the strongest operational signals in the current filtered view.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        for insight in current_insights:
            st.markdown(
                f"""
                <div class="gm-insight">
                    <div class="gm-insight-headline">{insight["headline"]}</div>
                    <div class="gm-insight-detail">{insight["detail"]}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    with signal_col:
        st.markdown(
            """
                <div class="gm-panel">
                    <div class="gm-panel-title">Signal Integrity Stack</div>
                    <div class="gm-panel-subtitle">This view keeps the new performance-bucket model while surfacing merge quality across trainer feedback, Top Brains, and both GitHub streams.</div>
                </div>
                """,
                unsafe_allow_html=True,
        )
        metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
        metric_col1.metric("Population", quality_metrics["Population"])
        metric_col2.metric("Feedback ID Match", quality_metrics["Feedback ID Matches"])
        metric_col3.metric("Top Brains ID Match", quality_metrics["Top Brains ID Matches"])
        metric_col4.metric("Assess GitHub ID Match", quality_metrics["Assessment GitHub ID Matches"])
        st.caption(
            f"Assignment GitHub ID match: {quality_metrics['Assignment GitHub ID Matches']} | "
            f"Name fallback: {quality_metrics['Name Fallbacks']} | "
            f"Missing assessment GitHub: {quality_metrics['Missing Assessment GitHub']} | "
            f"Missing assignment GitHub: {quality_metrics['Missing Assignment GitHub']} | "
            f"Average overall score: {pd.to_numeric(filtered_df['Overall Performance Score'], errors='coerce').mean():.1f}"
        )

    st.markdown(
        """
        <div class="gm-panel">
            <div class="gm-panel-title">Performance Bucket Dashboard</div>
            <div class="gm-panel-subtitle">Y-axis shows the performance score that drives the bucket. X-axis shows the final overall weighted score for management context.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    fig = px.scatter(
        filtered_df,
        x="Overall Performance Score",
        y="Performance Score",
        size="Bubble Size",
        color="Quadrant",
        color_discrete_map=QUADRANT_COLORS,
        size_max=28,
        custom_data=[
            "Superset ID",
            "Candidate Name",
            "Overall Performance Score",
            "Overall Grade",
            "Performance Score",
            "Performance Grade",
            "Assessment Composite Score",
            "Assessment Composite Grade",
            "Assessment GitHub Score",
            "Top Brains Score",
            "Assignment GitHub Score",
            "Attendance %",
            "Attendance Grade",
            "Trainer Feedback Score",
            "Quadrant",
            "College",
            "Assigned Batch",
            "Feedback Match Type",
            "Top Brains Match Type",
            "Assessment GitHub Match Type",
            "Assignment GitHub Match Type",
            "Assessment GitHub Weeks",
            "Top Brains Weeks",
            "Assignment GitHub Weeks",
        ],
    )
    fig.add_shape(type="rect", x0=0, y0=80, x1=100, y1=100, fillcolor=QUADRANT_COLORS["Deployable Candidates"], opacity=0.06, layer="below", line_width=0)
    fig.add_shape(type="rect", x0=0, y0=70, x1=100, y1=80, fillcolor=QUADRANT_COLORS["Progressing Candidates"], opacity=0.05, layer="below", line_width=0)
    fig.add_shape(type="rect", x0=0, y0=60, x1=100, y1=70, fillcolor=QUADRANT_COLORS["Basic Competency"], opacity=0.05, layer="below", line_width=0)
    fig.add_shape(type="rect", x0=0, y0=0, x1=100, y1=60, fillcolor=QUADRANT_COLORS["Critical Intervention"], opacity=0.05, layer="below", line_width=0)
    fig.add_hline(y=80, line_width=2, line_dash="dash", line_color="#94a3b8")
    fig.add_hline(y=70, line_width=2, line_dash="dash", line_color="#94a3b8")
    fig.add_hline(y=60, line_width=2, line_dash="dash", line_color="#94a3b8")
    fig.add_annotation(x=84, y=94, text="Deployable", showarrow=False, font=dict(size=22, color=QUADRANT_COLORS["Deployable Candidates"]))
    fig.add_annotation(x=84, y=75, text="Progressing", showarrow=False, font=dict(size=21, color=QUADRANT_COLORS["Progressing Candidates"]))
    fig.add_annotation(x=84, y=65, text="Basic<br>Competency", showarrow=False, font=dict(size=19, color=QUADRANT_COLORS["Basic Competency"]))
    fig.add_annotation(x=84, y=30, text="Critical<br>Intervention", showarrow=False, font=dict(size=20, color=QUADRANT_COLORS["Critical Intervention"]))
    fig.update_traces(
        marker=dict(line=dict(width=1, color="rgba(15,23,42,0.18)")),
        hovertemplate=(
            "<b>%{customdata[1]}</b><br>"
            "Superset ID: %{customdata[0]}<br>"
            "Overall Score: %{customdata[2]:.1f}<br>"
            "Overall Grade: %{customdata[3]}<br>"
            "Performance Score: %{customdata[4]:.1f}<br>"
            "Performance Grade: %{customdata[5]}<br>"
            "Assessment Composite: %{customdata[6]:.1f}<br>"
            "Assessment Composite Grade: %{customdata[7]}<br>"
            "Assessment GitHub: %{customdata[8]:.1f}<br>"
            "Top Brains: %{customdata[9]:.1f}<br>"
            "Assignment GitHub: %{customdata[10]:.1f}<br>"
            "Attendance: %{customdata[11]:.1f}%<br>"
            "Attendance Grade: %{customdata[12]}<br>"
            "Trainer Feedback Score: %{customdata[13]:.1f}<br>"
            "Bucket: %{customdata[14]}<br>"
            "College: %{customdata[15]}<br>"
            "Batch: %{customdata[16]}<br>"
            "Feedback Match: %{customdata[17]}<br>"
            "Top Brains Match: %{customdata[18]}<br>"
            "Assessment GitHub Match: %{customdata[19]}<br>"
            "Assignment GitHub Match: %{customdata[20]}<br>"
            "Assessment GitHub Weeks: %{customdata[21]}<br>"
            "Top Brains Weeks: %{customdata[22]}<br>"
            "Assignment GitHub Weeks: %{customdata[23]}<extra></extra>"
        ),
    )
    fig.update_layout(
        height=680,
        margin=dict(t=24, b=24, l=24, r=24),
        plot_bgcolor="white",
        paper_bgcolor="white",
        legend=dict(orientation="h", y=-0.16, x=0, title=""),
    )
    fig.update_xaxes(title="Overall Performance Score (50/35/10/5)", range=[0, 100], showgrid=False, zeroline=False)
    fig.update_yaxes(title="Performance Score (Assessment Composite + Assignments)", range=[0, 100], showgrid=False, zeroline=False)
    st.plotly_chart(fig, use_container_width=True)

    roster_source_df = filtered_df.copy()
    if current_quadrant_filter != "All Quadrants":
        roster_source_df = roster_source_df[roster_source_df["Quadrant"] == current_quadrant_filter].copy()
    roster_df = build_action_roster(roster_source_df)
    roster_display_df = build_action_roster_display(roster_source_df)
    st.markdown(
        """
        <div class="gm-panel">
            <div class="gm-panel-title">Exportable Executive Roster</div>
            <div class="gm-panel-subtitle">Filtered student roster for leadership reviews, trainer governance calls, and side-by-side comparison against the legacy Task-4 view.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.caption(f"Current cohort view: {current_quadrant_filter}")
    with st.expander("Column Guide", expanded=False):
        st.caption(
            "Superset ID = unique student ID | Standard_Status = current student status | "
            "Attendance % / Grade = attendance signal | Trainer Feedback Score = trainer rating converted to 100 | "
            "Assessment GitHub Score = average of GitHub assessment cycles | Top Brains Score = average Top Brains score | "
            "Assessment Composite Score = average of Assessment GitHub and Top Brains | "
            "Assignment GitHub Score = average of GitHub assignment cycles | "
            "Performance Score = weighted 50% assessment composite + 35% assignment | "
            "Overall Performance Score = final 50/35/10/5 score | "
            "Quadrant = performance-score bucket only: A+/A Deployable, B Progressing, C Basic Competency, F Critical Intervention."
        )
    dashboard_report_bytes = build_dashboard_report_bytes(
        roster_df,
        filtered_df,
        quality_metrics,
        current_insights,
        selected_scope,
        selected_college,
        selected_batch,
        current_quadrant_filter,
        eval_df,
        att_data,
    )
    export_col1, export_col2, export_col3, export_col4 = st.columns(4)
    with export_col1:
        st.download_button(
            "Download CSV",
            data=roster_df.to_csv(index=False).encode("utf-8"),
            file_name="magic_quadrant_plus_roster.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with export_col2:
        st.download_button(
            "Download Excel",
            data=to_excel_bytes(roster_df),
            file_name="magic_quadrant_plus_roster.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with export_col3:
        st.download_button(
            "Download Dashboard Report",
            data=dashboard_report_bytes,
            file_name="magic_quadrant_plus_dashboard_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with export_col4:
        if st.button("Publish Client Dashboard", use_container_width=True):
            published_path = publish_dashboard_report_file(dashboard_report_bytes)
            push_ok, push_message = push_published_dashboard_report()
            if push_ok:
                st.success(f"Published client dashboard workbook to {published_path}")
                st.info(push_message)
            else:
                st.warning(f"Published client dashboard workbook locally to {published_path}")
                st.error(push_message)
    st.dataframe(roster_display_df.fillna("N/A"), use_container_width=True, hide_index=True)
